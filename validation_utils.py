#!/usr/bin/env python3
"""
Validation Utilities for SEDO TSS Converter Pipeline
Provides robust input validation, error handling, and fallback mechanisms
"""

import openpyxl
import logging
from pathlib import Path
from typing import Union, Optional, List, Tuple, Dict
import re
import os
from difflib import SequenceMatcher

logger = logging.getLogger(__name__)

class ValidationError(Exception):
    """Custom exception for validation errors with actionable messages"""
    pass

class FileValidator:
    """
    Comprehensive file validation utilities
    Validates Excel files before processing to prevent pipeline failures
    """
    
    SUPPORTED_EXTENSIONS = {'.xlsx', '.xls', '.xlsm'}
    MAX_FILE_SIZE_MB = 100
    MIN_ROWS = 10
    MIN_COLS = 5
    
    @classmethod
    def validate_input_file(cls, file_path: Union[str, Path]) -> Path:
        """
        Comprehensive input file validation
        
        Args:
            file_path: Path to input file
            
        Returns:
            Validated Path object
            
        Raises:
            ValidationError: If validation fails with actionable message
        """
        file_path = Path(file_path)
        
        # Check file existence
        if not file_path.exists():
            raise ValidationError(
                f"Input file not found: {file_path}\n"
                f"Please check the file path and ensure the file exists."
            )
        
        # Check file extension
        if file_path.suffix.lower() not in cls.SUPPORTED_EXTENSIONS:
            raise ValidationError(
                f"Unsupported file format: {file_path.suffix}\n"
                f"Supported formats: {', '.join(cls.SUPPORTED_EXTENSIONS)}\n"
                f"Please convert your file to Excel format (.xlsx recommended)."
            )
        
        # Check file size
        file_size_mb = file_path.stat().st_size / (1024 * 1024)
        if file_size_mb > cls.MAX_FILE_SIZE_MB:
            logger.warning(
                f"⚠️  Large file detected: {file_size_mb:.1f}MB (recommended max: {cls.MAX_FILE_SIZE_MB}MB)\n"
                f"Processing may be slower. Consider splitting large files if possible."
            )
        
        # Check file permissions
        if not os.access(file_path, os.R_OK):
            raise ValidationError(
                f"Cannot read file: {file_path}\n"
                f"Please check file permissions and ensure the file is not locked by another application."
            )
        
        return file_path
    
    @classmethod
    def validate_excel_structure(cls, file_path: Path) -> Dict[str, any]:
        """
        Validate Excel file structure and content
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with file statistics
            
        Raises:
            ValidationError: If structure validation fails
        """
        try:
            # Try to load workbook
            workbook = openpyxl.load_workbook(str(file_path), read_only=True)
            worksheet = workbook.active
            
            # Get basic statistics
            stats = {
                'max_row': worksheet.max_row,
                'max_col': worksheet.max_column,
                'sheet_name': worksheet.title,
                'file_size_mb': file_path.stat().st_size / (1024 * 1024)
            }
            
            # Check minimum dimensions
            if stats['max_row'] < cls.MIN_ROWS:
                logger.warning(
                    f"⚠️  Small file detected: {stats['max_row']} rows "
                    f"(recommended min: {cls.MIN_ROWS} rows)"
                )
            
            if stats['max_col'] < cls.MIN_COLS:
                logger.warning(
                    f"⚠️  Narrow file detected: {stats['max_col']} columns "
                    f"(recommended min: {cls.MIN_COLS} columns)"
                )
            
            # Check for completely empty file
            has_data = False
            for row in range(1, min(11, stats['max_row'] + 1)):  # Check first 10 rows
                for col in range(1, min(11, stats['max_col'] + 1)):  # Check first 10 cols
                    if worksheet.cell(row, col).value is not None:
                        has_data = True
                        break
                if has_data:
                    break
            
            if not has_data:
                raise ValidationError(
                    f"Empty Excel file detected: {file_path}\n"
                    f"Please ensure the file contains data in the first 10 rows and columns."
                )
            
            workbook.close()
            logger.debug(f"Excel validation passed: {stats}")
            return stats
            
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise ValidationError(
                f"Invalid Excel file: {file_path}\n"
                f"Error: {str(e)}\n"
                f"Please ensure this is a valid Excel file and not corrupted."
            )
        except PermissionError as e:
            raise ValidationError(
                f"Permission denied accessing file: {file_path}\n"
                f"Please close the file in Excel and try again."
            )
        except Exception as e:
            raise ValidationError(
                f"Failed to read Excel file: {file_path}\n"
                f"Error: {str(e)}\n"
                f"Please check if the file is corrupted or in use by another application."
            )

class HeaderDetector:
    """
    Robust header detection with fuzzy matching and fallbacks
    Handles variations in header text and positioning
    """
    
    @classmethod
    def similarity(cls, a: str, b: str) -> float:
        """Calculate similarity between two strings (0-1)"""
        return SequenceMatcher(None, a.lower(), b.lower()).ratio()
    
    @classmethod
    def find_header_fuzzy(cls, worksheet, target_patterns: List[str], 
                         min_similarity: float = 0.7, max_search_rows: int = 50) -> Optional[Tuple[int, int, str, float]]:
        """
        Find header using fuzzy matching with multiple patterns
        
        Args:
            worksheet: openpyxl worksheet
            target_patterns: List of patterns to search for
            min_similarity: Minimum similarity score (0-1)
            max_search_rows: Maximum rows to search
            
        Returns:
            Tuple of (row, col, matched_text, similarity) or None if not found
        """
        best_match = None
        best_score = 0
        
        for row in range(1, min(max_search_rows + 1, worksheet.max_row + 1)):
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row, col).value
                if not cell_value or not isinstance(cell_value, str):
                    continue
                
                cell_text = cell_value.strip()
                if not cell_text:
                    continue
                
                for pattern in target_patterns:
                    similarity = cls.similarity(cell_text, pattern)
                    if similarity >= min_similarity and similarity > best_score:
                        best_score = similarity
                        best_match = (row, col, cell_text, similarity)
                        logger.debug(f"Header match: '{cell_text}' ~ '{pattern}' (similarity: {similarity:.2f})")
        
        return best_match
    
    @classmethod
    def find_general_type_header(cls, worksheet) -> Optional[Tuple[int, int, str]]:
        """
        Find 'General Type/Sub-Type in Connect' header with fallbacks
        
        Returns:
            Tuple of (row, col, matched_text) or None if not found
        """
        # Primary patterns (exact matches first)
        primary_patterns = [
            "General Type/Sub-Type in Connect",
            "General Type of Material in Connect"
        ]
        
        # Fallback patterns (fuzzy matches)
        fallback_patterns = [
            "general type sub-type connect",
            "general type material connect", 
            "general type connect",
            "sub-type connect",
            "material connect",
            "general type",
            "sub type"
        ]
        
        # Try primary patterns with exact matching
        for pattern in primary_patterns:
            for row in range(1, min(51, worksheet.max_row + 1)):
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row, col).value
                    if cell_value and isinstance(cell_value, str):
                        if pattern.lower() in cell_value.lower().strip():
                            logger.debug(f"Found exact header match: '{cell_value}' at row {row}, col {col}")
                            return (row, col, cell_value)
        
        # Try fallback patterns with fuzzy matching
        result = cls.find_header_fuzzy(worksheet, fallback_patterns, min_similarity=0.6)
        if result:
            row, col, matched_text, similarity = result
            logger.info(f"Found header using fallback: '{matched_text}' (similarity: {similarity:.2f})")
            return (row, col, matched_text)
        
        return None
    
    @classmethod
    def find_article_headers(cls, worksheet, max_search_rows: int = 20) -> Optional[Tuple[int, int, int]]:
        """
        Find Article Name and Article Number headers with fallbacks
        
        Returns:
            Tuple of (name_col, number_col, header_row) or None if not found
        """
        name_patterns = ["Article Name", "article name", "name", "product name"]
        number_patterns = ["Article No.", "Article No", "article no", "number", "product no", "art no"]
        
        for row in range(1, min(max_search_rows + 1, worksheet.max_row + 1)):
            name_col = None
            number_col = None
            
            # Scan row for article headers
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row, col).value
                if not cell_value or not isinstance(cell_value, str):
                    continue
                
                cell_text = cell_value.strip().lower()
                
                # Check for name patterns
                if not name_col and any(pattern.lower() in cell_text for pattern in name_patterns):
                    name_col = col
                    logger.debug(f"Found Article Name at row {row}, col {col}: '{cell_value}'")
                
                # Check for number patterns  
                if not number_col and any(pattern.lower() in cell_text for pattern in number_patterns):
                    number_col = col
                    logger.debug(f"Found Article Number at row {row}, col {col}: '{cell_value}'")
            
            # If both found in same row, return result
            if name_col and number_col:
                return (name_col, number_col, row)
        
        return None

class ErrorHandler:
    """
    Enhanced error handling with specific exceptions and recovery suggestions
    """
    
    @classmethod
    def handle_file_error(cls, error: Exception, file_path: Path, operation: str) -> str:
        """
        Generate actionable error message for file operations
        
        Args:
            error: The original exception
            file_path: Path that caused the error
            operation: Description of operation (e.g., "loading", "saving")
            
        Returns:
            Formatted error message with suggestions
        """
        error_type = type(error).__name__
        
        if isinstance(error, FileNotFoundError):
            return (
                f"File not found during {operation}: {file_path}\n"
                f"Suggestions:\n"
                f"- Check if the file path is correct\n"
                f"- Ensure the file hasn't been moved or deleted\n"
                f"- Verify you have permission to access the directory"
            )
        
        elif isinstance(error, PermissionError):
            return (
                f"Permission denied during {operation}: {file_path}\n"
                f"Suggestions:\n"
                f"- Close the file if it's open in Excel or another application\n"
                f"- Check if you have read/write permissions\n"
                f"- Try running as administrator if necessary"
            )
        
        elif "InvalidFileException" in error_type:
            return (
                f"Invalid Excel file during {operation}: {file_path}\n"
                f"Suggestions:\n"
                f"- Ensure the file is a valid Excel format (.xlsx, .xls, .xlsm)\n"
                f"- Try opening the file in Excel to check for corruption\n"
                f"- Re-save the file in Excel format if it was converted"
            )
        
        else:
            return (
                f"Unexpected error during {operation}: {file_path}\n"
                f"Error type: {error_type}\n"
                f"Error details: {str(error)}\n"
                f"Suggestions:\n"
                f"- Check if the file is corrupted\n"
                f"- Ensure sufficient disk space\n"
                f"- Try with a different file to isolate the issue"
            )
    
    @classmethod
    def handle_header_not_found(cls, header_type: str, searched_patterns: List[str], max_rows_searched: int) -> str:
        """
        Generate helpful message when required headers are not found
        """
        return (
            f"Required header not found: {header_type}\n"
            f"Searched for patterns: {', '.join(searched_patterns)}\n"
            f"Search area: first {max_rows_searched} rows\n"
            f"Suggestions:\n"
            f"- Check if your Excel file has the expected header structure\n"
            f"- Verify the header text matches one of the expected patterns\n"
            f"- Ensure headers are not in merged cells (run Step 1 first)\n"
            f"- Try with the -v flag for detailed search information"
        )

def validate_pipeline_input(file_path: Union[str, Path], step_name: str) -> Path:
    """
    Comprehensive pipeline input validation
    
    Args:
        file_path: Input file path
        step_name: Name of the step (for error context)
        
    Returns:
        Validated Path object
        
    Raises:
        ValidationError: If validation fails
    """
    logger.debug(f"Validating input for {step_name}: {file_path}")
    
    try:
        # Basic file validation
        validated_path = FileValidator.validate_input_file(file_path)
        
        # Excel structure validation
        stats = FileValidator.validate_excel_structure(validated_path)
        
        # Log file info
        logger.info(f"📊 File stats: {stats['max_row']} rows, {stats['max_col']} cols, {stats['file_size_mb']:.1f}MB")
        
        return validated_path
        
    except ValidationError:
        # Re-raise validation errors as-is (they have good messages)
        raise
    except Exception as e:
        # Wrap unexpected errors
        raise ValidationError(
            f"Unexpected validation error in {step_name}:\n"
            f"{str(e)}\n"
            f"Please check your input file and try again."
        )