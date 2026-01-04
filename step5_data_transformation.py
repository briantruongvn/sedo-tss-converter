#!/usr/bin/env python3
"""
Step 5: Data Transformation - Content Population
SEDO TSS Converter Pipeline Step 5/6

LOGIC:
- Transform processed data from Step 2 into populated template from Step 4
- Maps source data (A,B,E,F,H) to base columns (B,C,D,F,P) with "TR" document type
- Scans horizontal data (columns J to "Oldest TR date") excluding N/A values
- Creates one output row per valid horizontal cell, expanding data vertically
- Starts data population from row 11 to preserve article information area

PIPELINE POSITION: Fifth step - populates template with actual compliance data
INPUT: Header-processed data (output-X-Step2.xlsx) + Article template (output-X-Step4.xlsx)
OUTPUT: Populated data file (output-X-Step5.xlsx)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from pathlib import Path
from typing import Union, Optional, List, Tuple
import argparse
import sys
import re
import shutil
from validation_utils import ValidationError, handle_validation_error
from pipeline_validator import validate_before_pipeline
from pipeline_config import PipelineConfig

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DataTransformer:
    """
    Data Transformer for Step 5 - Content Population
    
    Advanced data transformation and expansion:
    - Combines processed headers from Step 2 with formatted template from Step 4
    - Maps base data (General Type, Sub-Type, Material, Producer, H-column) to output columns (Level)
    - Scans and expands horizontal test requirements into individual rows
    - Filters out N/A values and preserves only valid compliance data
    - Creates structured database-ready format starting from row 11
    """
    
    @classmethod
    def get_metadata(cls):
        """Get step metadata from centralized configuration"""
        return PipelineConfig.get_step(5)
    
    @property
    def step_name(self):
        """Get step display name"""
        return self.get_metadata().display_name
    
    @property
    def step_description(self):
        """Get step description"""
        return self.get_metadata().description
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "data" / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def transform_data(self, step2_file: Union[str, Path], 
                      step4_file: Union[str, Path],
                      output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Transform data from Step 2 to Step 5 using Step 4 template
        
        Args:
            step2_file: Data source file (output-X-Step2.xlsx)
            step4_file: Template file (output-X-Step4.xlsx)
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to transformed file
        """
        logger.info("📋 Step 5: Data Transformation - Content Population")
        
        step2_path = Path(step2_file)
        step4_path = Path(step4_file)
        
        if not step2_path.exists():
            raise FileNotFoundError(f"Step 2 file not found: {step2_path}")
        if not step4_path.exists():
            raise FileNotFoundError(f"Step 4 file not found: {step4_path}")
        
        # Auto-generate output file if not provided
        if output_file is None:
            file_num = self._extract_file_number(step2_path.name)
            if file_num:
                output_file = self.output_dir / f"output-{file_num}-Step5.xlsx"
            else:
                base_name = step2_path.stem.replace('-Step2', '')
                output_file = self.output_dir / f"{base_name}-Step5.xlsx"
        else:
            output_file = Path(output_file)
        
        logger.info(f"Data Source: {step2_path}")
        logger.info(f"Template: {step4_path}")
        logger.info(f"Output: {output_file}")
        
        # Load source data and template
        source_wb = openpyxl.load_workbook(str(step2_path))
        source_ws = source_wb.active
        
        # Copy template as starting point
        shutil.copy2(str(step4_path), str(output_file))
        output_wb = openpyxl.load_workbook(str(output_file))
        output_ws = output_wb.active
        
        # Step 1: Find header row
        header_row = self._find_header_row(source_ws)
        if header_row is None:
            raise ValueError("Could not find 'General Type/Sub-Type in Connect' header row in source file")
        
        requirements_row = header_row + 4
        first_data_row = header_row + 6
        logger.info(f"Found header at row {header_row}, requirements at row {requirements_row}, data starts at row {first_data_row}")
        
        # Step 2: Process all data rows
        rows_created = self._process_all_data_rows(source_ws, output_ws, header_row, first_data_row)
        
        logger.info(f"✅ Created {rows_created} data rows")
        
        # Save result
        try:
            output_wb.save(str(output_file))
            logger.info(f"✅ Step 5 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        return str(output_file)
    
    def _find_header_row(self, worksheet) -> Optional[int]:
        """Find row containing 'General Type/Sub-Type in Connect' header (case-insensitive)"""
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Search for "General Type/Sub-Type in Connect" header
        target_text = "general type/sub-type in connect"
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if target_text in cell_value.strip().lower():
                        logger.debug(f"Found 'General Type/Sub-Type in Connect' header at row {row}, column {chr(64+col)}")
                        return row
        
        return None
    
    def _find_oldest_tr_date_column(self, worksheet, header_row: int) -> Optional[int]:
        """Find column containing 'Oldest TR date' in the header row"""
        max_col = worksheet.max_column
        
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(header_row, col).value
            if cell_value and isinstance(cell_value, str):
                if "oldest tr date" in cell_value.lower().strip():
                    logger.debug(f"Found 'Oldest TR date' at column {chr(64+col)} (column {col})")
                    return col
        
        return None
    
    def _extract_base_data(self, worksheet, data_row: int) -> dict:
        """
        Extract base mapping data from data row
        
        Mappings:
        - A{data_row} → B4 (General Type Component)
        - B{data_row} → C4 (Sub-Type Component) 
        - E{data_row} → F4 (Producer)
        - F{data_row} → D4 (Material Designation)
        - H{data_row} → P4 (Level)
        - "TR" → H4 (Document type)
        """
        return {
            'general_type': worksheet.cell(data_row, 1).value,  # A → B4
            'sub_type': worksheet.cell(data_row, 2).value,      # B → C4
            'producer': worksheet.cell(data_row, 5).value,      # E → F4  
            'material_designation': worksheet.cell(data_row, 6).value,  # F → D4
            'h_value': worksheet.cell(data_row, 8).value,       # H → P4
            'document_type': 'TR'  # Hardcoded → H4
        }
    
    def _scan_horizontal_data(self, worksheet, data_row: int, header_row: int) -> List[int]:
        """
        Scan horizontally from column J to column before 'Oldest TR date'
        Skips: empty cells, "N/A" values, whitespace-only cells
        
        Returns list of column numbers that have valid data
        """
        valid_cols = []
        
        # Find 'Oldest TR date' column to limit scan range
        oldest_tr_col = self._find_oldest_tr_date_column(worksheet, header_row)
        
        # Set end column: if 'Oldest TR date' found, stop before it; otherwise use max column
        if oldest_tr_col and oldest_tr_col > 10:
            end_col = oldest_tr_col - 1
            logger.debug(f"Found 'Oldest TR date' at column {chr(64+oldest_tr_col)}, limiting scan to column {chr(64+end_col)}")
        else:
            end_col = worksheet.max_column
            logger.debug(f"'Oldest TR date' not found, scanning to max column {chr(64+end_col)}")
        
        # Start from column J (10) and scan to end_col
        for col in range(10, end_col + 1):
            cell_value = worksheet.cell(data_row, col).value
            if self._is_valid_cell_value(cell_value):
                valid_cols.append(col)
        
        logger.debug(f"Valid columns (excluding N/A): {[chr(64+col) for col in valid_cols]}")
        return valid_cols
    
    def _is_valid_cell_value(self, cell_value) -> bool:
        """
        Check if cell value is valid for processing
        
        Returns False for:
        - None/empty values
        - "N/A" (case-insensitive)
        - Whitespace-only strings
        
        Returns True for all other values
        """
        if not cell_value:
            return False
        
        cell_str = str(cell_value).strip()
        
        # Skip empty strings after stripping
        if not cell_str:
            return False
        
        # Skip N/A values (case-insensitive)
        if cell_str.lower() == 'n/a':
            return False
        
        return True
    
    def _extract_four_row_data(self, worksheet, col: int, header_row: int) -> dict:
        """
        Extract 5-row data for a specific column using header row + offsets:
        - header_row + 0 → Column I (Requirement source)
        - header_row + 1 → Column J (Sub-type detail)
        - header_row + 2 → Column K (Regulation)  
        - header_row + 4 → Column N (Frequency)
        - header_row + 5 → Column L (Limit)
        """
        return {
            'requirement_source': worksheet.cell(header_row, col).value,     # header_row + 0 → I
            'sub_type_detail': worksheet.cell(header_row + 1, col).value,    # header_row + 1 → J
            'regulation': worksheet.cell(header_row + 2, col).value,         # header_row + 2 → K
            'frequency': worksheet.cell(header_row + 4, col).value,          # header_row + 4 → N
            'limit': worksheet.cell(header_row + 5, col).value,              # header_row + 5 → L
        }
    
    def _process_all_data_rows(self, source_ws, output_ws, header_row: int, first_data_row: int) -> int:
        """
        Process all data rows from first_data_row until no more data
        
        Returns total number of output rows created
        """
        current_output_row = 11  # Start populating output from row 11
        total_rows_created = 0
        
        # Find last row with data
        last_data_row = self._find_last_data_row(source_ws, first_data_row)
        logger.info(f"Processing data rows {first_data_row} to {last_data_row}")
        
        # Process each data row
        for data_row in range(first_data_row, last_data_row + 1):
            # Extract base data for this row
            base_data = self._extract_base_data(source_ws, data_row)
            
            # Scan horizontal data for this row
            horizontal_cols = self._scan_horizontal_data(source_ws, data_row, header_row)
            
            if not horizontal_cols:
                logger.debug(f"No horizontal data found in row {data_row}, skipping")
                continue
            
            logger.debug(f"Row {data_row}: Base data = {base_data}, Horizontal cols = {len(horizontal_cols)}")
            
            # Process each non-empty horizontal cell
            for col_idx, source_col in enumerate(horizontal_cols):
                # Extract 4-row data for this column
                four_row_data = self._extract_four_row_data(source_ws, source_col, header_row)
                
                # Copy A:H data from previous row if not first output row
                if current_output_row > 11:
                    self._copy_base_formatting(output_ws, 11, current_output_row)
                
                # Populate base data (A-H columns)
                self._populate_base_data(output_ws, current_output_row, base_data)
                
                # Populate 5-column data (I-L, N)
                output_ws.cell(current_output_row, 9, four_row_data['requirement_source'])   # I
                output_ws.cell(current_output_row, 10, four_row_data['sub_type_detail'])     # J
                output_ws.cell(current_output_row, 11, four_row_data['regulation'])          # K
                output_ws.cell(current_output_row, 12, four_row_data['limit'])               # L
                output_ws.cell(current_output_row, 14, four_row_data['frequency'])           # N
                
                logger.debug(f"Output row {current_output_row}: Data row {data_row}, Col {chr(64+source_col)} → {four_row_data}")
                current_output_row += 1
                total_rows_created += 1
        
        return total_rows_created
    
    def _find_last_data_row(self, worksheet, first_data_row: int) -> int:
        """Find the last row that contains data in columns A or B"""
        max_row = worksheet.max_row
        
        # Search backwards from max_row to find last row with data in A or B
        for row in range(max_row, first_data_row - 1, -1):
            for col in range(1, 3):  # Columns A-B
                cell_value = worksheet.cell(row, col).value
                if cell_value and str(cell_value).strip():
                    return row
        
        return first_data_row  # Fallback to first_data_row if no data found
    
    def _copy_base_formatting(self, output_ws, source_row: int, target_row: int):
        """Copy formatting from source row to target row for columns A-H"""
        for col in range(1, 9):  # Columns A-H
            source_cell = output_ws.cell(source_row, col)
            target_cell = output_ws.cell(target_row, col)
            
            # Copy value and formatting
            target_cell.value = source_cell.value
            if source_cell.font:
                target_cell.font = Font(
                    bold=source_cell.font.bold,
                    color=source_cell.font.color
                )
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color,
                    fill_type=source_cell.fill.fill_type
                )
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )
    
    def _populate_base_data(self, output_ws, row: int, base_data: dict):
        """Populate base data in columns A-H and P"""
        output_ws.cell(row, 2, base_data['general_type'])         # B
        output_ws.cell(row, 3, base_data['sub_type'])             # C
        output_ws.cell(row, 4, base_data['material_designation']) # D
        output_ws.cell(row, 6, base_data['producer'])             # F
        output_ws.cell(row, 8, base_data['document_type'])        # H
        output_ws.cell(row, 16, base_data['h_value'])             # P (column 16)
    
    def _extract_file_number(self, filename: str) -> str:
        """Extract file number from filename like 'output-1-Step2.xlsx'"""
        match = re.search(r'output-(\d+)', filename)
        return match.group(1) if match else ""
    
    def transform_multiple_files(self, step2_patterns: list, output_dir: Optional[str] = None) -> list:
        """
        Transform multiple files matching patterns
        
        Args:
            step2_patterns: List of Step 2 file patterns or paths
            output_dir: Output directory (if None, use default)
            
        Returns:
            List of output file paths
        """
        if output_dir:
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(parents=True, exist_ok=True)
        
        results = []
        
        for pattern in step2_patterns:
            # Handle glob patterns
            if '*' in str(pattern):
                step2_files = list(self.base_dir.glob(str(pattern)))
            else:
                step2_files = [Path(pattern)]
            
            for step2_file in step2_files:
                if step2_file.exists() and step2_file.suffix.lower() in ['.xlsx', '.xls']:
                    try:
                        # Find corresponding Step 4 file
                        file_num = self._extract_file_number(step2_file.name)
                        if file_num:
                            step4_file = self.base_dir / "data" / "output" / f"output-{file_num}-Step4.xlsx"
                        else:
                            base_name = step2_file.stem.replace('-Step2', '')
                            step4_file = self.base_dir / "data" / "output" / f"{base_name}-Step4.xlsx"
                        
                        if not step4_file.exists():
                            logger.error(f"❌ Step 4 template not found: {step4_file}")
                            continue
                        
                        result = self.transform_data(step2_file, step4_file)
                        results.append(result)
                        logger.info(f"✅ Processed: {step2_file} + {step4_file} → {result}")
                    except Exception as e:
                        logger.error(f"❌ Failed to process {step2_file}: {e}")
                else:
                    logger.warning(f"⚠️  Skipped: {step2_file} (not found or not Excel file)")
        
        return results

def main():
    """Command line interface for standalone data transformation"""
    parser = argparse.ArgumentParser(description='Data Transformer Step 5 - Standalone')
    parser.add_argument('input', nargs='+', help='Input file(s) or patterns (output-X-Step2.xlsx)')
    parser.add_argument('-o', '--output', help='Output file or directory')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    parser.add_argument('--batch', action='store_true', help='Batch mode for multiple files')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize transformer
    transformer = DataTransformer(args.base_dir)
    
    try:
        if args.batch or len(args.input) > 1:
            # Multiple files mode
            output_dir = args.output if args.output else None
            results = transformer.transform_multiple_files(args.input, output_dir)
            
            print("\n📊 Batch Processing Results:")
            print(f"✅ Successfully processed: {len(results)} files")
            for result in results:
                print(f"   📁 {result}")
                
        else:
            # Single file mode - need both Step2 and Step4 files
            step2_file = args.input[0]
            
            # Auto-find Step4 file
            file_num = transformer._extract_file_number(Path(step2_file).name)
            if file_num:
                step4_file = transformer.base_dir / "data" / "output" / f"output-{file_num}-Step4.xlsx"
            else:
                base_name = Path(step2_file).stem.replace('-Step2', '')
                step4_file = transformer.base_dir / "data" / "output" / f"{base_name}-Step4.xlsx"
            
            if not step4_file.exists():
                print(f"❌ Error: Step 4 template not found: {step4_file}")
                sys.exit(1)
            
            result = transformer.transform_data(step2_file, step4_file, args.output)
            print(f"\n✅ Success!")
            print(f"📁 Output: {result}")
            
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()