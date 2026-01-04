#!/usr/bin/env python3
"""
Step 7: Finished Product Processing - Data Cleanup
SEDO TSS Converter Pipeline Step 7/8

LOGIC:
- Process "Finished product" entries for data cleanup
- Scan column B from row 11 downward for finished product patterns
- Apply transformations: A="Art", B/C/D/F=empty for matched rows
- Clean and standardize product classification data

PIPELINE POSITION: Seventh step - data cleanup before final processing
INPUT: Complete processed file (data/output/output-X-Step6.xlsx)
OUTPUT: Cleaned data file (data/output/output-X-Step7.xlsx)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from pathlib import Path
from typing import Union, Optional, List, Tuple
import argparse
import sys
import re
import shutil
from validation_utils import ValidationError, handle_validation_error
from pipeline_validator import validate_before_pipeline
from pipeline_config import PipelineConfig

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class FinishedProductProcessor:
    """
    Finished Product Processor for Step 7 - Data Cleanup
    
    Processes finished product entries for data standardization:
    - Scans column B from row 11 downward for finished product patterns
    - Applies cleanup transformations to matching rows
    - Converts "finished product" entries to standardized "Art" format
    - Clears redundant data in specified columns for cleaner output
    """
    
    @classmethod
    def get_metadata(cls):
        """Get step metadata from centralized configuration"""
        return PipelineConfig.get_step(7)
    
    @property
    def step_name(self):
        """Get step display name"""
        return self.get_metadata().display_name
    
    @property
    def step_description(self):
        """Get step description"""
        return self.get_metadata().description
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "data" / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def process_finished_products(self, step6_file: Union[str, Path],
                                  output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process finished product entries for data cleanup
        
        Args:
            step6_file: Input file from Step 6 (output-X-Step6.xlsx)
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to processed file
        """
        logger.info("📋 Step 7: Finished Product Processing - Data Cleanup")
        
        step6_path = Path(step6_file)
        if not step6_path.exists():
            raise FileNotFoundError(f"Step 6 file not found: {step6_path}")
        
        # Auto-generate output file if not provided
        if output_file is None:
            file_num = self._extract_file_number(step6_path.name)
            if file_num:
                output_file = self.output_dir / f"output-{file_num}-Step7.xlsx"
            else:
                base_name = step6_path.stem.replace('-Step6', '')
                output_file = self.output_dir / f"{base_name}-Step7.xlsx"
        else:
            output_file = Path(output_file)
        
        logger.info(f"Input (Step 6): {step6_path}")
        logger.info(f"Output: {output_file}")
        
        # Copy Step 6 as starting point
        shutil.copy2(str(step6_path), str(output_file))
        
        # Load copied file for processing
        try:
            wb = openpyxl.load_workbook(str(output_file))
            ws = wb.active
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            raise
        
        # Step 7.1: Process finished product rows
        processed_rows = self._process_finished_product_rows(ws)
        logger.info(f"✅ Step 7.1: Processed {processed_rows} finished product rows")
        
        # Step 7.2: Process article matching
        matched_rows = self._process_article_matching(ws)
        logger.info(f"✅ Step 7.2: Processed {matched_rows} article matching rows")
        
        # Save result
        try:
            wb.save(str(output_file))
            logger.info(f"✅ Step 7 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        return str(output_file)
    
    def _process_finished_product_rows(self, worksheet) -> int:
        """
        Process finished product rows starting from row 11
        
        Logic:
        - Scan column B from B11 downward
        - Check for "finished product", "finish product", "finish" (case insensitive)
        - For matching rows: A="Art", B/C/D/F=empty
        
        Returns:
            Number of rows processed
        """
        processed_count = 0
        max_row = worksheet.max_row
        
        logger.debug(f"Scanning column B from row 11 to {max_row}")
        
        for row in range(11, max_row + 1):
            # Get value from column B
            b_cell = worksheet.cell(row, 2)  # Column B
            b_value = b_cell.value
            
            # Check if this is a finished product row
            if self._is_finished_product(b_value):
                logger.debug(f"Row {row}: Found finished product pattern: '{b_value}'")
                
                # Apply transformations
                worksheet.cell(row, 1).value = "Art"  # A = "Art"
                worksheet.cell(row, 2).value = ""     # B = empty
                worksheet.cell(row, 3).value = ""     # C = empty  
                worksheet.cell(row, 4).value = ""     # D = empty
                worksheet.cell(row, 6).value = ""     # F = empty
                
                processed_count += 1
                logger.debug(f"Row {row}: Applied transformations - A='Art', B/C/D/F=empty")
        
        return processed_count
    
    def _is_finished_product(self, value) -> bool:
        """
        Check if value matches finished product patterns
        
        Patterns (case insensitive):
        - "finished product"
        - "finish product" 
        - "finish"
        
        Returns:
            True if value matches any pattern, False otherwise
        """
        if not value or not isinstance(value, str):
            return False
        
        value_lower = value.lower().strip()
        
        # Define patterns to match
        patterns = [
            "finished product",
            "finish product", 
            "finish"
        ]
        
        # Check if any pattern is contained in the value
        for pattern in patterns:
            if pattern in value_lower:
                return True
        
        return False
    
    def _get_article_headers(self, worksheet) -> List[Tuple[int, str]]:
        """
        Get Article Names from headers starting at R1
        
        Returns:
            List of (column_number, article_name) tuples
        """
        article_headers = []
        max_col = worksheet.max_column
        
        # Start from column R (18) and scan rightward
        for col in range(18, max_col + 1):
            cell_value = worksheet.cell(1, col).value
            if cell_value and isinstance(cell_value, str) and cell_value.strip():
                article_name = cell_value.strip()
                article_headers.append((col, article_name))
                logger.debug(f"Found Article Name at column {chr(64+col) if col <= 26 else f'col{col}'}: '{article_name}'")
            else:
                # Stop at first empty column (articles should be consecutive)
                break
        
        logger.debug(f"Found {len(article_headers)} article headers")
        return article_headers
    
    def _parse_multiline_p_value(self, value) -> List[str]:
        """
        Parse multi-line value from P column
        
        Returns:
            List of cleaned lines
        """
        if not value:
            return []
        
        value_str = str(value).strip()
        if not value_str:
            return []
        
        # Split by newlines and filter out empty lines
        lines = [line.strip() for line in value_str.split('\n') if line.strip()]
        
        return lines
    
    def _match_article_name(self, p_line: str, article_name: str) -> bool:
        """
        Check if P line matches Article Name
        
        Uses substring matching with case-insensitive comparison
        
        Args:
            p_line: Line from P column value
            article_name: Article name from header
            
        Returns:
            True if P line contains Article Name, False otherwise
        """
        if not p_line or not article_name:
            return False
        
        # Clean both values for comparison
        p_clean = p_line.lower().strip()
        article_clean = article_name.lower().strip()
        
        # Check if article name is contained in P line
        return article_clean in p_clean
    
    def _is_all_items(self, p_value) -> bool:
        """
        Check if P value contains patterns indicating all items
        
        Patterns (case insensitive):
        - "All"
        - "All items"  
        - "All products"
        
        Args:
            p_value: Value from P column
            
        Returns:
            True if P contains any "All" pattern, False otherwise
        """
        if not p_value:
            return False
        
        p_str = str(p_value).lower().strip()
        if not p_str:
            return False
        
        # Define "All" patterns
        all_patterns = [
            "all",
            "all items",
            "all products"
        ]
        
        # Check if any pattern is contained in P value
        for pattern in all_patterns:
            if pattern in p_str:
                return True
        
        return False
    
    def _process_article_matching(self, worksheet) -> int:
        """
        Process article matching for P column starting from P11
        
        Logic:
        - Scan P column from P11 downward
        - Parse multi-line values
        - Match each line with Article Names from headers
        - Fill "X" in corresponding article columns
        - If P empty, fill "X" in all article columns
        
        Returns:
            Number of rows processed
        """
        # Get article headers
        article_headers = self._get_article_headers(worksheet)
        if not article_headers:
            logger.warning("No article headers found - skipping article matching")
            return 0
        
        processed_count = 0
        max_row = worksheet.max_row
        
        logger.debug(f"Processing article matching from P11 to P{max_row}")
        logger.debug(f"Article columns: {[(chr(64+col) if col <= 26 else f'col{col}', name) for col, name in article_headers]}")
        
        for row in range(11, max_row + 1):
            # Get value from column P
            p_cell = worksheet.cell(row, 16)  # Column P = 16
            p_value = p_cell.value
            
            # Parse P value (could be multi-line)
            p_lines = self._parse_multiline_p_value(p_value)
            
            if not p_lines:
                # P is empty - fill "X" in all article columns
                for col, article_name in article_headers:
                    worksheet.cell(row, col).value = "X"
                logger.debug(f"Row {row}: P empty - filled all article columns with 'X'")
                processed_count += 1
                continue
            
            # Check if P contains "All" patterns - fill all article columns
            if self._is_all_items(p_value):
                for col, article_name in article_headers:
                    worksheet.cell(row, col).value = "X"
                logger.debug(f"Row {row}: P contains 'All' pattern - filled all article columns with 'X'")
                processed_count += 1
                continue
            
            # Track which articles matched
            matched_articles = set()
            
            # Check each P line against each article
            for p_line in p_lines:
                for col, article_name in article_headers:
                    if self._match_article_name(p_line, article_name):
                        matched_articles.add(col)
                        logger.debug(f"Row {row}: '{p_line}' matches '{article_name}' -> column {chr(64+col) if col <= 26 else f'col{col}'}")
            
            # Fill "X" for matched articles
            if matched_articles:
                for col in matched_articles:
                    worksheet.cell(row, col).value = "X"
                logger.debug(f"Row {row}: Filled 'X' in columns {[chr(64+col) if col <= 26 else f'col{col}' for col in matched_articles]}")
            else:
                # No matches found - log for debugging
                logger.debug(f"Row {row}: No article matches found for P value: {p_lines}")
            
            processed_count += 1
        
        return processed_count
    
    def _extract_file_number(self, filename: str) -> str:
        """Extract file number from filename like 'output-1-Step6.xlsx'"""
        match = re.search(r'output-(\d+)', filename)
        return match.group(1) if match else ""
    
    def process_multiple_files(self, step6_patterns: list, output_dir: Optional[str] = None) -> list:
        """
        Process multiple Step 6 files
        
        Args:
            step6_patterns: List of Step 6 file patterns or paths
            output_dir: Output directory (if None, use default)
            
        Returns:
            List of output file paths
        """
        if output_dir:
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(parents=True, exist_ok=True)
        
        results = []
        
        for pattern in step6_patterns:
            # Handle glob patterns
            if '*' in str(pattern):
                step6_files = list(self.base_dir.glob(str(pattern)))
            else:
                step6_files = [Path(pattern)]
            
            for step6_file in step6_files:
                if step6_file.exists() and step6_file.suffix.lower() in ['.xlsx', '.xls']:
                    try:
                        result = self.process_finished_products(step6_file)
                        results.append(result)
                        logger.info(f"✅ Processed: {step6_file} → {result}")
                    except Exception as e:
                        logger.error(f"❌ Failed to process {step6_file}: {e}")
                else:
                    logger.warning(f"⚠️  Skipped: {step6_file} (not found or not Excel file)")
        
        return results

def main():
    """Command line interface for standalone finished product processing"""
    parser = argparse.ArgumentParser(description='Finished Product Processor Step 7 - Standalone')
    parser.add_argument('input', nargs='*', help='Input Step 6 file(s) or patterns (output-X-Step6.xlsx). If not provided, uses data/output/*-Step6.xlsx')
    parser.add_argument('-o', '--output', help='Output file or directory')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    parser.add_argument('--batch', action='store_true', help='Batch mode for multiple files')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize processor
    processor = FinishedProductProcessor(args.base_dir)
    
    try:
        if args.batch or len(args.input) > 1 or not args.input:
            # Multiple files mode or auto-detect mode
            input_patterns = args.input if args.input else ["data/output/*-Step6.xlsx"]
            output_dir = args.output if args.output else None
            results = processor.process_multiple_files(input_patterns, output_dir)
            
            print("\n📊 Batch Processing Results:")
            print(f"✅ Successfully processed: {len(results)} files")
            for result in results:
                print(f"   📁 {result}")
                
        else:
            # Single file mode
            step6_file = args.input[0]
            output_file = args.output
            
            result = processor.process_finished_products(step6_file, output_file)
            print(f"\n✅ Success!")
            print(f"📁 Output: {result}")
            
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()