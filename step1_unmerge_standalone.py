#!/usr/bin/env python3
"""
Step 1: Excel Cell Unmerging - Foundation of Pipeline
SEDO TSS Converter Pipeline Step 1/6

LOGIC:
- Unmerge all merged cell ranges in Excel file
- Preserve data by filling empty cells with merged cell values
- Maintain formatting and structure while making cells accessible

PIPELINE POSITION: First step - prepares file for header processing
INPUT: Raw Excel file with merged cells (data/input/input-X.xlsx)
OUTPUT: Unmerged Excel file (data/output/output-X-Step1.xlsx)
"""

import openpyxl
import logging
from pathlib import Path
from typing import Union, Optional
import argparse
import sys
import re
from validation_utils import validate_pipeline_input, ValidationError, ErrorHandler

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelUnmerger:
    """
    Standalone Excel Cell Unmerger
    
    Simple and effective approach:
    - Unmerge all merged cells without discrimination
    - Fill each cell with its original top-left value
    - No complex analysis, just reliable data preservation
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "data" / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def unmerge_file(self, input_file: Union[str, Path], 
                    output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Simple & Effective Cell Unmerging
        
        Args:
            input_file: Input Excel file path
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to unmerged file
        """
        logger.info("📋 Excel Cell Unmerging - Standalone")
        
        # Comprehensive input validation
        input_path = validate_pipeline_input(input_file, "Step 1")
        
        # Auto-generate output file if not provided
        if output_file is None:
            file_num = self._extract_file_number(input_path.name)
            if file_num:
                output_file = self.output_dir / f"output-{file_num}-Step1.xlsx"
            else:
                base_name = input_path.stem
                output_file = self.output_dir / f"{base_name}-Step1.xlsx"
        else:
            output_file = Path(output_file)
        
        logger.info(f"Input: {input_path}")
        logger.info(f"Output: {output_file}")
        
        # Load workbook with enhanced error handling
        try:
            wb = openpyxl.load_workbook(str(input_path))
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, input_path, "loading workbook")
            logger.error(error_msg)
            raise ValidationError(error_msg)
        
        ws = wb.active
        
        # Get all merged ranges
        merged_ranges = list(ws.merged_cells.ranges)
        total_merges = len(merged_ranges)
        
        if total_merges == 0:
            logger.info("No merged cells found - saving file as-is")
            wb.save(str(output_file))
            return str(output_file)
        
        logger.info(f"Found {total_merges} merged ranges to process")
        
        # Step 1: Efficient value collection
        value_map = {}
        processed_cells = 0
        
        for merge_range in merged_ranges:
            min_col, min_row, max_col, max_row = merge_range.bounds
            top_left_value = ws.cell(min_row, min_col).value
            
            # Map all cells in range to the top-left value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    value_map[(row, col)] = top_left_value
                    processed_cells += 1
        
        logger.info(f"Collected values for {processed_cells} cells from {total_merges} ranges")
        
        # Step 2: Unmerge all ranges
        unmerged_count = 0
        failed_unmerges = []
        
        for merge_range in merged_ranges:
            try:
                ws.unmerge_cells(str(merge_range))
                unmerged_count += 1
            except Exception as e:
                failed_unmerges.append((merge_range, str(e)))
                logger.warning(f"Failed to unmerge {merge_range}: {e}")
        
        logger.info(f"✅ Unmerged {unmerged_count}/{total_merges} ranges successfully")
        if failed_unmerges:
            logger.warning(f"❌ {len(failed_unmerges)} ranges failed to unmerge")
        
        # Step 3: Fill all cells with preserved values
        filled_count = 0
        for (row, col), value in value_map.items():
            cell = ws.cell(row, col)
            if cell.value is None and value is not None:
                cell.value = value
                filled_count += 1
        
        logger.info(f"✅ Filled {filled_count} empty cells with preserved values")
        
        # Calculate and log efficiency
        fill_rate = (filled_count / processed_cells) * 100 if processed_cells > 0 else 0
        logger.info(f"📊 Data preservation: {fill_rate:.1f}% ({filled_count}/{processed_cells} cells)")
        
        # Save result with enhanced error handling
        try:
            wb.save(str(output_file))
            logger.info(f"✅ Unmerge completed: {output_file}")
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, Path(output_file), "saving workbook")
            logger.error(error_msg)
            raise ValidationError(error_msg)
        
        return str(output_file)
    
    def _extract_file_number(self, filename: str) -> str:
        """Extract file number from filename like 'input-1.xlsx' or 'Input-1.xlsx'"""
        match = re.search(r'[Ii]nput-(\d+)', filename)
        return match.group(1) if match else ""
    
    def unmerge_multiple_files(self, input_patterns: list, output_dir: Optional[str] = None) -> list:
        """
        Unmerge multiple files matching patterns
        
        Args:
            input_patterns: List of file patterns or paths
            output_dir: Output directory (if None, use default)
            
        Returns:
            List of output file paths
        """
        if output_dir:
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(parents=True, exist_ok=True)
        
        results = []
        
        for pattern in input_patterns:
            # Handle glob patterns
            if '*' in str(pattern):
                input_files = list(self.base_dir.glob(str(pattern)))
            else:
                input_files = [Path(pattern)]
            
            for input_file in input_files:
                if input_file.exists() and input_file.suffix.lower() in ['.xlsx', '.xls']:
                    try:
                        result = self.unmerge_file(input_file)
                        results.append(result)
                        logger.info(f"✅ Processed: {input_file} → {result}")
                    except Exception as e:
                        logger.error(f"❌ Failed to process {input_file}: {e}")
                else:
                    logger.warning(f"⚠️  Skipped: {input_file} (not found or not Excel file)")
        
        return results

def main():
    """Command line interface for standalone unmerging"""
    parser = argparse.ArgumentParser(description='Excel Cell Unmerger - Standalone')
    parser.add_argument('input', nargs='+', help='Input file(s) or patterns')
    parser.add_argument('-o', '--output', help='Output file or directory')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    parser.add_argument('--batch', action='store_true', help='Batch mode for multiple files')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize unmerger
    unmerger = ExcelUnmerger(args.base_dir)
    
    try:
        if args.batch or len(args.input) > 1:
            # Multiple files mode
            output_dir = args.output if args.output else None
            results = unmerger.unmerge_multiple_files(args.input, output_dir)
            
            print("\n📊 Batch Processing Results:")
            print(f"✅ Successfully processed: {len(results)} files")
            for result in results:
                print(f"   📁 {result}")
                
        else:
            # Single file mode
            input_file = args.input[0]
            output_file = args.output
            
            result = unmerger.unmerge_file(input_file, output_file)
            print(f"\n✅ Success!")
            print(f"📁 Output: {result}")
            
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()