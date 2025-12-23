#!/usr/bin/env python3
"""
Debug Step 4 boundary detection to understand why Lovetex gets included
"""

import openpyxl
import sys
from pathlib import Path

def debug_step4_boundary():
    """Debug Step 4 boundary detection"""
    
    step2_file = Path("data/output/Test Summary of CIRKUSTÄLT chld tent red blue-white 2025_UPDATE (system)-Step2.xlsx")
    
    if not step2_file.exists():
        print(f"❌ Step2 file not found: {step2_file}")
        return
    
    print(f"🔍 Debugging Step 4 boundary detection in: {step2_file}")
    
    # Load workbook
    wb = openpyxl.load_workbook(str(step2_file))
    ws = wb.active
    
    # Find first data row (after header+2)
    header_row = None
    for row in range(1, 30):
        for col in range(1, 15):
            cell_value = ws.cell(row, col).value
            if cell_value and "general type" in str(cell_value).lower():
                header_row = row
                print(f"Found 'General Type' at row {row}, col {chr(64+col)}")
                break
        if header_row:
            break
    
    if not header_row:
        print("❌ Could not find header row with 'General Type'")
        return
        
    first_data_row = header_row + 3
    print(f"Header row: {header_row}")
    print(f"First data row: {first_data_row}")
    
    # Check last row detection logic
    max_row = ws.max_row
    print(f"Max row in worksheet: {max_row}")
    
    # Search backwards from max_row to find last row with data in A or B
    print(f"\n🔍 Checking columns A-B for last data row:")
    found_last_row = None
    
    for row in range(max_row, first_data_row - 1, -1):
        for col in range(1, 3):  # Columns A-B
            cell_value = ws.cell(row, col).value
            if cell_value and str(cell_value).strip():
                print(f"   Row {row}, Col {chr(64+col)}: '{cell_value}'")
                if not found_last_row:
                    found_last_row = row
                    print(f"   ⭐ LAST DATA ROW DETECTED: {row}")
                break
        
        # Stop after finding first 5 rows with data
        if found_last_row and row < found_last_row - 5:
            break
    
    print(f"\n📊 BOUNDARY ANALYSIS:")
    print(f"First data row: {first_data_row}")
    print(f"Last data row: {found_last_row}")
    
    # Check rows around the expected boundary (row 149)
    print(f"\n🔍 Checking rows around 149:")
    for row in range(145, 155):
        a_val = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value
        f_val = ws.cell(row, 6).value
        
        a_str = str(a_val).strip() if a_val else ""
        b_str = str(b_val).strip() if b_val else ""
        f_str = str(f_val).strip() if f_val else ""
        
        status = ""
        if (a_str and a_str) or (b_str and b_str):
            status = "⚠️ HAS A/B DATA"
        
        print(f"   Row {row}: A='{a_str}' B='{b_str}' F='{f_str}' {status}")
    
    # Check where Lovetex appears
    print(f"\n🔍 Checking for Lovetex in Step 2 source:")
    lovetex_rows = []
    for row in range(1, max_row + 1):
        for col in range(1, 15):
            cell_value = ws.cell(row, col).value
            if cell_value and 'lovetex' in str(cell_value).lower():
                lovetex_rows.append((row, col, cell_value))
    
    if lovetex_rows:
        print(f"Found {len(lovetex_rows)} Lovetex entries in Step 2:")
        for row, col, value in lovetex_rows[:10]:  # Show first 10
            print(f"   Row {row}, Col {chr(64+col)}: '{value}'")
    else:
        print("No Lovetex found in Step 2 source")
    
    wb.close()

if __name__ == "__main__":
    debug_step4_boundary()