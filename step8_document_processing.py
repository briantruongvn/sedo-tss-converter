#!/usr/bin/env python3
"""
Step 8: Document Processing - Final Optimization
SEDO TSS Converter Pipeline Step 8/8

LOGIC:
- Remove "finished product" rows to clean data
- Fill document type and requirement source specifications
- Apply final data optimization and cleanup rules
- Produce clean, production-ready output

PIPELINE POSITION: Eighth step - final processing and optimization
INPUT: Complete processed file (data/output/output-X-Step6.xlsx)
OUTPUT: Final optimized file (data/output/output-X-Step8.xlsx)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from pathlib import Path
from typing import Union, Optional, List, Tuple
import argparse
import sys
import re
import shutil
from validation_utils import ValidationError, handle_validation_error
from pipeline_validator import validate_before_pipeline
from pipeline_config import PipelineConfig

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DocumentProcessor:
    """
    Document Processor for Step 8 - Final Optimization
    
    Processes Step 7 output with three main tasks:
    8.1 Remove rows containing "finished product" in column Q
    8.2 Fill document type (column H) and requirement source (column I) based on column Q
    8.3 Clear column P from P11 onwards for final cleanup
    """
    
    @classmethod
    def get_metadata(cls):
        """Get step metadata from centralized configuration"""
        return PipelineConfig.get_step(8)
    
    @property
    def step_name(self):
        """Get step display name"""
        return self.get_metadata().display_name
    
    @property
    def step_description(self):
        """Get step description"""
        return self.get_metadata().description
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "data" / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def process_step7(self, step6_file: Union[str, Path], 
                      output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process Step 7 file: remove finished product rows, fill document specs, and clear column P
        
        Args:
            step6_file: Step 6 output file (output-X-Step6.xlsx)
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to processed file
        """
        logger.info("📋 Step 8: Document Processing - Final Optimization")
        
        step6_path = Path(step6_file)
        
        if not step6_path.exists():
            raise FileNotFoundError(f"Step 6 file not found: {step6_path}")
        
        # Auto-generate output file if not provided
        if output_file is None:
            file_num = self._extract_file_number(step6_path.name)
            if file_num:
                output_file = self.output_dir / f"output-{file_num}-Step8.xlsx"
            else:
                # Handle both Step6 and Step7 input files
                base_name = step6_path.stem.replace('-Step6', '').replace('-Step7', '')
                output_file = self.output_dir / f"{base_name}-Step8.xlsx"
        else:
            output_file = Path(output_file)
        
        logger.info(f"Input: {step6_path}")
        logger.info(f"Output: {output_file}")
        
        # Copy Step 6 as starting point
        shutil.copy2(str(step6_path), str(output_file))
        
        # Load workbook
        wb = openpyxl.load_workbook(str(output_file))
        ws = wb.active
        
        logger.info(f"Original rows: {ws.max_row}")
        
        # Step 8.1: Remove "finished product" rows
        removed_count = self._remove_finished_product_rows(ws)
        logger.info(f"✅ Removed {removed_count} 'finished product' rows")
        
        # Step 8.2: Fill document type and requirement source
        filled_count = self._fill_document_specs(ws)
        logger.info(f"✅ Filled document specs for {filled_count} rows")
        
        # Step 8.3: Clear column P from P11 to end
        cleared_count = self._clear_column_p(ws)
        logger.info(f"✅ Cleared {cleared_count} cells in column P (P11 onwards)")
        
        logger.info(f"Final rows: {ws.max_row}")
        
        # Save result
        try:
            wb.save(str(output_file))
            logger.info(f"✅ Step 8 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        return str(output_file)
    
    def _remove_finished_product_rows(self, worksheet) -> int:
        """
        Remove rows containing 'finished product' in column Q (case-insensitive)
        
        Returns:
            Number of rows removed
        """
        logger.debug("🔄 Scanning for 'finished product' rows in column Q...")
        
        rows_to_delete = []
        
        # Scan all rows for "finished product" in column Q (column 17) - skip header rows
        for row in range(11, worksheet.max_row + 1):  # Start from row 11 (after headers)
            q_value = worksheet.cell(row, 17).value  # Column Q = 17
            if q_value and isinstance(q_value, str):
                if 'finished product' in q_value.lower():
                    rows_to_delete.append(row)
                    logger.debug(f"Marking row {row} for deletion: '{q_value}'")
        
        # Delete rows in reverse order to maintain indices
        for row_num in sorted(rows_to_delete, reverse=True):
            worksheet.delete_rows(row_num, 1)
            logger.debug(f"Deleted row {row_num}")
        
        return len(rows_to_delete)
    
    def _fill_document_specs(self, worksheet) -> int:
        """
        Fill document type (column H) and requirement source (column I) based on column Q

        Logic:
        - Document Type: First word (no spaces) from column Q → column H
          * SKIP if Q starts with "N/" pattern (e.g., "1/", "2/", "3/")
        - Requirement Source: Extract IOS/MAT patterns from column Q → column I

        Returns:
            Number of rows processed
        """
        logger.debug("🔄 Filling document type and requirement source...")
        
        rows_processed = 0
        
        # Process all data rows (skip header rows 1-10)
        for row in range(11, worksheet.max_row + 1):  # Start from row 11 (after headers)
            q_value = worksheet.cell(row, 17).value  # Column Q = 17
            
            # Skip empty or non-string values
            if not q_value or not isinstance(q_value, str):
                continue
            
            q_text = q_value.strip()
            if not q_text:
                continue
            
            # Parse document information
            doc_type, req_source = self._parse_document_info(q_text)

            # Fill document type in column H (column 8)
            # IMPORTANT: Do NOT overwrite "SD" values from Step 6
            existing_h = worksheet.cell(row, 8).value
            if doc_type and existing_h != "SD":  # Only fill if NOT an SD row
                worksheet.cell(row, 8, doc_type)
                logger.debug(f"Row {row}: Document type '{doc_type}' → H{row}")
            
            # Fill requirement source in column I (column 9)
            if req_source:
                worksheet.cell(row, 9, req_source)
                logger.debug(f"Row {row}: Requirement source '{req_source}' → I{row}")
            
            rows_processed += 1
        
        return rows_processed
    
    def _parse_document_info(self, q_text: str) -> Tuple[str, str]:
        """
        Parse document type and requirement source from column Q text

        Args:
            q_text: Text from column Q

        Returns:
            (document_type, requirement_source) tuple
        """
        # Check if Q text starts with "N/" pattern (number followed by slash)
        # If so, skip document type extraction (leave column H empty)
        # Examples: "1/", "2/", "3/" → don't fill column H
        if re.match(r'^\d+/', q_text.strip()):
            doc_type = ""  # Don't fill column H for "1/", "2/", "3/" patterns
        else:
            # Extract document type (first word without spaces)
            words = q_text.split()
            doc_type = words[0] if words else ""

        # Extract requirement sources containing IOS or MAT
        req_sources = self._extract_requirement_sources(q_text)
        req_source = " & ".join(req_sources) if req_sources else ""

        return doc_type, req_source
    
    def _extract_requirement_sources(self, text: str) -> List[str]:
        """
        Extract requirement source patterns containing IOS or MAT
        
        Enhanced patterns:
        - MAT patterns: MAT0250, MAT-0250, MAT0250: Jiangsu → MAT0250
        - IOS patterns: IOS-PRG-0272, IOS-MAT-0010 → IOS-PRG-0272
        
        Args:
            text: Text to search for patterns
            
        Returns:
            List of matched requirement source patterns
        """
        requirement_sources = []
        
        # Split text by common separators to handle multiple entries
        segments = re.split(r'[&,;]+', text)
        
        for segment in segments:
            segment = segment.strip()
            if not segment:
                continue
            
            # Look for complex IOS pattern first (to avoid double matching with MAT)
            # IOS with prefix: IOS-MAT-0010, IOS-PRG-0272, IOS- PRG-0273 (with spaces)
            ios_complex_pattern = r'\bIOS-\s*[A-Z]{2,4}-\d+'
            ios_complex_matches = re.findall(ios_complex_pattern, segment, re.IGNORECASE)
            
            # Look for simple IOS pattern: IOS-0123
            ios_simple_pattern = r'\bIOS-\d+'
            ios_simple_matches = re.findall(ios_simple_pattern, segment, re.IGNORECASE)
            
            # Look for MAT pattern only if no IOS-MAT found
            segment_temp = segment
            for ios_match in ios_complex_matches:
                segment_temp = segment_temp.replace(ios_match, '')
            
            mat_pattern = r'\bMAT[-]?\d+(?=[\s:]|$)'
            mat_matches = re.findall(mat_pattern, segment_temp, re.IGNORECASE)
            
            # Add all matches from this segment
            for match in ios_complex_matches + ios_simple_matches + mat_matches:
                requirement_sources.append(match.upper())
        
        # Remove duplicates while preserving order
        seen = set()
        unique_sources = []
        for source in requirement_sources:
            if source not in seen:
                seen.add(source)
                unique_sources.append(source)
        
        return unique_sources
    
    def _clear_column_p(self, worksheet) -> int:
        """
        Clear column P from P11 to end of data
        
        This cleans up the P column after we've used it for article matching
        in Step 7, keeping the data clean for final output.
        
        Returns:
            Number of cells cleared
        """
        logger.debug("🔄 Clearing column P from P11 onwards...")
        
        cleared_count = 0
        max_row = worksheet.max_row
        
        # Clear column P (column 16) from row 11 onwards
        for row in range(11, max_row + 1):
            p_cell = worksheet.cell(row, 16)  # Column P = 16
            if p_cell.value is not None:
                p_cell.value = None
                cleared_count += 1
                logger.debug(f"Cleared P{row}")
        
        return cleared_count
    
    def _extract_file_number(self, filename: str) -> str:
        """Extract file number from filename like 'output-1-Step6.xlsx'"""
        match = re.search(r'output-(\d+)', filename)
        return match.group(1) if match else ""
    
    def process_multiple_files(self, step6_patterns: list, output_dir: Optional[str] = None) -> list:
        """
        Process multiple Step 6 files
        
        Args:
            step6_patterns: List of Step 6 file patterns or paths
            output_dir: Output directory (if None, use default)
            
        Returns:
            List of output file paths
        """
        if output_dir:
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(parents=True, exist_ok=True)
        
        results = []
        
        for pattern in step6_patterns:
            # Handle glob patterns
            if '*' in str(pattern):
                step6_files = list(self.base_dir.glob(str(pattern)))
            else:
                step6_files = [Path(pattern)]
            
            for step6_file in step6_files:
                if step6_file.exists() and step6_file.suffix.lower() in ['.xlsx', '.xls']:
                    try:
                        result = self.process_step7(step6_file)
                        results.append(result)
                        logger.info(f"✅ Processed: {step6_file} → {result}")
                    except Exception as e:
                        logger.error(f"❌ Failed to process {step6_file}: {e}")
                else:
                    logger.warning(f"⚠️  Skipped: {step6_file} (not found or not Excel file)")
        
        return results

def main():
    """Command line interface for standalone document processing"""
    parser = argparse.ArgumentParser(description='Document Processor Step 8 - Standalone')
    parser.add_argument('input', nargs='*', help='Input Step 6 file(s) or patterns (output-X-Step6.xlsx). If not provided, uses data/output/*-Step6.xlsx')
    parser.add_argument('-o', '--output', help='Output file or directory')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    parser.add_argument('--batch', action='store_true', help='Batch mode for multiple files')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize processor
    processor = DocumentProcessor(args.base_dir)
    
    try:
        if args.batch or len(args.input) > 1 or not args.input:
            # Multiple files mode or auto-detect mode
            input_patterns = args.input if args.input else ["data/output/*-Step6.xlsx"]
            output_dir = args.output if args.output else None
            results = processor.process_multiple_files(input_patterns, output_dir)
            
            print("\n📊 Batch Processing Results:")
            print(f"✅ Successfully processed: {len(results)} files")
            for result in results:
                print(f"   📁 {result}")
                
        else:
            # Single file mode
            step6_file = args.input[0]
            result = processor.process_step7(step6_file, args.output)
            print(f"\n✅ Success!")
            print(f"📁 Output: {result}")
            
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()