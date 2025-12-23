#!/usr/bin/env python3
"""
Step 2: Header Processing with 3-Case Logic
SEDO TSS Converter Pipeline Step 2/6

LOGIC:
- Find "General Type/Sub-Type in Connect" header row (n)
- Process 3 rows below header (n+1, n+2, n+3) with 3-case logic:
  * Case 1: All same → empty, keep middle, empty
  * Case 2: First≠Second=Third → keep first, keep second, empty  
  * Case 3: All different → keep first, combine second+third, empty

PIPELINE POSITION: Second step - processes headers after unmerging
INPUT: Unmerged file (data/output/output-X-Step1.xlsx)
OUTPUT: Header-processed file (data/output/output-X-Step2.xlsx)
"""

import openpyxl
import logging
from pathlib import Path
from typing import Union, Optional, Tuple
import argparse
import sys
import re
from validation_utils import validate_pipeline_input, ValidationError, ErrorHandler, HeaderDetector

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class HeaderProcessor:
    """
    Standalone Header Processor for Step 2
    
    Processes 3 rows after "General Type/Sub-Type in Connect" header row with specific 3-case logic:
    - Case 1: val16==val17==val18 → empty, keep val17, empty
    - Case 2: val16!=val17==val18 → keep val16, keep val17, empty  
    - Case 3: val16!=val17!=val18 → keep val16, val17+" "+val18, empty
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "data" / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def process_file(self, input_file: Union[str, Path], 
                    output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process header rows with 3-case logic
        
        Args:
            input_file: Input file from Step 1 (output-X-Step1.xlsx)
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to processed file
        """
        logger.info("📋 Step 2: Header Processing with 3-Case Logic")
        
        # Comprehensive input validation
        input_path = validate_pipeline_input(input_file, "Step 2")
        
        # Auto-generate output file if not provided
        if output_file is None:
            file_num = self._extract_file_number(input_path.name)
            if file_num:
                output_file = self.output_dir / f"output-{file_num}-Step2.xlsx"
            else:
                base_name = input_path.stem.replace('-Step1', '')
                output_file = self.output_dir / f"{base_name}-Step2.xlsx"
        else:
            output_file = Path(output_file)
        
        logger.info(f"Input: {input_path}")
        logger.info(f"Output: {output_file}")
        
        # Load workbook with enhanced error handling
        try:
            wb = openpyxl.load_workbook(str(input_path))
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, input_path, "loading workbook")
            logger.error(error_msg)
            raise ValidationError(error_msg)
        
        ws = wb.active
        
        # Step 1: Find header row with enhanced detection
        header_result = HeaderDetector.find_general_type_header(ws)
        if header_result is None:
            searched_patterns = ["General Type/Sub-Type in Connect", "General Type of Material in Connect"]
            error_msg = ErrorHandler.handle_header_not_found("General Type header", searched_patterns, 50)
            logger.error(error_msg)
            raise ValidationError(error_msg)
        
        header_row, header_col, matched_text = header_result
        logger.info(f"Found header: '{matched_text}' at row {header_row}, column {chr(64+header_col)}")
        
        logger.info(f"Found header row: {header_row}")
        
        # Step 2: Identify processing range
        start_col = 10  # Column J
        last_data_col = self._find_last_data_column(ws, header_row)
        
        logger.info(f"Processing columns {chr(64+start_col)} to {chr(64+last_data_col)} (columns {start_col}-{last_data_col})")
        logger.info(f"Processing rows {header_row+1} to {header_row+3}")
        
        # Step 3: Apply 3-case logic to all columns
        processed_columns = self._process_header_columns(ws, header_row, start_col, last_data_col)
        
        logger.info(f"✅ Processed {processed_columns} columns with 3-case logic")
        
        # Save result
        try:
            wb.save(str(output_file))
            logger.info(f"✅ Step 2 completed: {output_file}")
        except Exception as e:
            error_msg = ErrorHandler.handle_file_error(e, Path(output_file), "saving workbook")
            logger.error(error_msg)
            raise ValidationError(error_msg)
        
        return str(output_file)
    
    def _find_header_row(self, worksheet) -> Optional[int]:
        """Find row containing 'General Type/Sub-Type in Connect' (case-insensitive)"""
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Search entire worksheet for "General Type/Sub-Type in Connect" or "General Type of Material in Connect"
        target_texts = ["general type/sub-type in connect", "general type of material in connect"]
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    cell_text = cell_value.strip().lower()
                    if any(target in cell_text for target in target_texts):
                        logger.debug(f"Found header text at row {row}, column {chr(64+col)}")
                        return row
        
        return None
    
    def _find_oldest_tr_date_column(self, worksheet, header_row: int) -> Optional[int]:
        """Find column containing 'Oldest TR date' in the header row (n)"""
        max_col = worksheet.max_column
        
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(header_row, col).value
            if cell_value and isinstance(cell_value, str):
                if "oldest tr date" in cell_value.lower().strip():
                    logger.debug(f"Found 'Oldest TR date' at column {chr(64+col)} (column {col})")
                    return col
        
        return None

    def _find_last_data_column(self, worksheet, header_row: int) -> int:
        """Find last column with actual data (column before 'Oldest TR date')"""
        # First try to find 'Oldest TR date' column
        oldest_tr_col = self._find_oldest_tr_date_column(worksheet, header_row)
        
        if oldest_tr_col and oldest_tr_col > 1:
            # Return column before 'Oldest TR date'
            last_col = oldest_tr_col - 1
            logger.debug(f"Last data column based on 'Oldest TR date': {chr(64+last_col)} (column {last_col})")
            return last_col
        
        # Fallback: Start from reasonable right boundary and work backwards
        for col in range(worksheet.max_column, 0, -1):
            # Check if any cell in this column has data
            for row in range(1, min(worksheet.max_row + 1, 50)):  # Check first 50 rows
                cell_value = worksheet.cell(row, col).value
                if cell_value and str(cell_value).strip():
                    logger.debug(f"Last data column found (fallback): {chr(64+col)} (column {col})")
                    return col
        
        # Final fallback to max_column if no data found
        return worksheet.max_column
    
    def _process_header_columns(self, worksheet, header_row: int, 
                               start_col: int, end_col: int) -> int:
        """
        Process all columns with 3-case logic
        
        Args:
            worksheet: Excel worksheet
            header_row: Row number containing "General Type/Sub-Type in Connect" (n)
            start_col: Start column (J = 10)
            end_col: End column with data
            
        Returns:
            Number of columns processed
        """
        processed_count = 0
        
        # Process each column from J to last data column
        for col in range(start_col, end_col + 1):
            # Get values from 3 rows: n+1, n+2, n+3
            val1 = worksheet.cell(header_row + 1, col).value  # Row n+1
            val2 = worksheet.cell(header_row + 2, col).value  # Row n+2  
            val3 = worksheet.cell(header_row + 3, col).value  # Row n+3
            
            # Normalize values for processing
            val1_str = self._normalize_value(val1)
            val2_str = self._normalize_value(val2)
            val3_str = self._normalize_value(val3)
            
            # Apply 3-case logic
            new_val1, new_val2, new_val3 = self._apply_three_case_logic(val1_str, val2_str, val3_str)
            
            # Write back to worksheet (only if changed)
            if new_val1 != val1_str:
                worksheet.cell(header_row + 1, col).value = new_val1 if new_val1 else None
            if new_val2 != val2_str:
                worksheet.cell(header_row + 2, col).value = new_val2 if new_val2 else None
            if new_val3 != val3_str:
                worksheet.cell(header_row + 3, col).value = new_val3 if new_val3 else None
            
            # Log changes for debugging (only for first few columns)
            if col <= start_col + 3:  # Log first 4 columns for debugging
                change_indicator = ""
                if (new_val1, new_val2, new_val3) != (val1_str, val2_str, val3_str):
                    change_indicator = " [CHANGED]"
                logger.debug(f"Column {chr(64+col)}: val16='{val1_str}' val17='{val2_str}' val18='{val3_str}' → val16='{new_val1}' val17='{new_val2}' val18='{new_val3}'{change_indicator}")
            
            processed_count += 1
        
        return processed_count
    
    def _normalize_value(self, value) -> str:
        """Normalize cell value to string for comparison"""
        if value is None:
            return ""
        elif isinstance(value, str):
            return value.strip()
        else:
            # Convert other types (datetime, numbers) to string
            return str(value).strip()
    
    def _apply_three_case_logic(self, val1: str, val2: str, val3: str) -> Tuple[str, str, str]:
        """
        Apply 3-case logic as specified:
        
        Case 1: val16 == val17 == val18 → empty, keep val17, empty
        Case 2: val16 != val17 && val17 == val18 → keep val16, keep val17, empty
        Case 3: val16 != val17 && val17 != val18 → keep val16, val17 + " " + val18, empty
        
        Args:
            val1, val2, val3: String values from the 3 rows
            
        Returns:
            Tuple of (new_val1, new_val2, new_val3)
        """
        # Normalize for comparison (but preserve original formatting for output)
        norm_val1 = val1.strip() if val1 else ""
        norm_val2 = val2.strip() if val2 else ""
        norm_val3 = val3.strip() if val3 else ""
        
        # Case 1: All three values are the same
        if norm_val1 == norm_val2 == norm_val3 and norm_val1 != "":
            return "", val2, ""
        
        # Case 2: First different from second, but second equals third
        elif norm_val1 != norm_val2 and norm_val2 == norm_val3:
            return val1, val2, ""
        
        # Case 3: All different OR first two same but different from third
        else:
            # Combine val2 and val3 with space, handle empty values
            if norm_val2 and norm_val3:
                combined_val2 = f"{val2} {val3}".strip()
            elif norm_val2:
                combined_val2 = val2
            elif norm_val3:
                combined_val2 = val3
            else:
                combined_val2 = ""
            
            return val1, combined_val2, ""
    
    def _extract_file_number(self, filename: str) -> str:
        """Extract file number from filename like 'output-1-Step1.xlsx'"""
        match = re.search(r'output-(\d+)', filename)
        return match.group(1) if match else ""
    
    def process_multiple_files(self, input_patterns: list, output_dir: Optional[str] = None) -> list:
        """
        Process multiple files matching patterns
        
        Args:
            input_patterns: List of file patterns or paths
            output_dir: Output directory (if None, use default)
            
        Returns:
            List of output file paths
        """
        if output_dir:
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(parents=True, exist_ok=True)
        
        results = []
        
        for pattern in input_patterns:
            # Handle glob patterns
            if '*' in str(pattern):
                input_files = list(self.base_dir.glob(str(pattern)))
            else:
                input_files = [Path(pattern)]
            
            for input_file in input_files:
                if input_file.exists() and input_file.suffix.lower() in ['.xlsx', '.xls']:
                    try:
                        result = self.process_file(input_file)
                        results.append(result)
                        logger.info(f"✅ Processed: {input_file} → {result}")
                    except Exception as e:
                        logger.error(f"❌ Failed to process {input_file}: {e}")
                else:
                    logger.warning(f"⚠️  Skipped: {input_file} (not found or not Excel file)")
        
        return results

def main():
    """Command line interface for standalone header processing"""
    parser = argparse.ArgumentParser(description='Header Processor Step 2 - Standalone')
    parser.add_argument('input', nargs='+', help='Input file(s) or patterns (output-X-Step1.xlsx)')
    parser.add_argument('-o', '--output', help='Output file or directory')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    parser.add_argument('--batch', action='store_true', help='Batch mode for multiple files')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize processor
    processor = HeaderProcessor(args.base_dir)
    
    try:
        if args.batch or len(args.input) > 1:
            # Multiple files mode
            output_dir = args.output if args.output else None
            results = processor.process_multiple_files(args.input, output_dir)
            
            print("\n📊 Batch Processing Results:")
            print(f"✅ Successfully processed: {len(results)} files")
            for result in results:
                print(f"   📁 {result}")
                
        else:
            # Single file mode
            input_file = args.input[0]
            output_file = args.output
            
            result = processor.process_file(input_file, output_file)
            print(f"\n✅ Success!")
            print(f"📁 Output: {result}")
            
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()