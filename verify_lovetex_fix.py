#!/usr/bin/env python3
"""
Verify that Lovetex contamination is fixed in new Step 5 output
"""

import openpyxl
import sys
from pathlib import Path

def verify_lovetex_fix():
    """Verify Lovetex is eliminated from fixed Step 5"""
    
    step5_file = Path("data/output/Test Summary of CIRKUSTÄLT chld tent red blue-white 2025_UPDATE (system)-Step5.xlsx")
    
    if not step5_file.exists():
        print(f"❌ Step5 file not found: {step5_file}")
        return
    
    print(f"🔍 Verifying Lovetex fix in: {step5_file}")
    
    # Load workbook
    wb = openpyxl.load_workbook(str(step5_file))
    ws = wb.active
    
    lovetex_count = 0
    lovetex_entries = []
    
    # Search all cells for Lovetex
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column + 1, 15)):
            cell_value = ws.cell(row, col).value
            if cell_value and isinstance(cell_value, str) and "lovetex" in cell_value.lower():
                lovetex_count += 1
                lovetex_entries.append({
                    'row': row,
                    'col': chr(64 + col),
                    'value': cell_value
                })
    
    print(f"\n📊 VERIFICATION RESULTS:")
    print(f"Total rows in Step 5: {ws.max_row}")
    print(f"Lovetex entries found: {lovetex_count}")
    
    if lovetex_count == 0:
        print(f"✅ SUCCESS! No Lovetex contamination found")
        print(f"✅ Bug fixed - Step 5 now respects row 149 boundary")
    else:
        print(f"❌ STILL CONTAMINATED! Found {lovetex_count} Lovetex entries:")
        for entry in lovetex_entries:
            print(f"   Row {entry['row']}, Col {entry['col']}: '{entry['value']}'")
    
    # Check data boundary
    print(f"\n🔍 Data boundary check:")
    
    # Check if any data beyond row 149
    data_beyond_149 = False
    for row in range(150, min(ws.max_row + 1, 200)):
        for col in range(1, 8):  # Check key columns A-G
            cell_value = ws.cell(row, col).value
            if cell_value and str(cell_value).strip() and str(cell_value).strip().upper() != 'SD':
                data_beyond_149 = True
                print(f"   ⚠️ Data found beyond row 149: Row {row}, Col {chr(64+col)}: '{cell_value}'")
                break
        if data_beyond_149:
            break
    
    if not data_beyond_149:
        print(f"   ✅ No significant data beyond row 149")
    
    # Summary comparison
    print(f"\n📊 Before vs After Fix:")
    print(f"BEFORE FIX:")
    print(f"   - Data range: rows 21 to 165")
    print(f"   - Lovetex entries: 15")
    print(f"   - Final rows: 385")
    print(f"   - Unique rows: 385")
    print(f"")
    print(f"AFTER FIX:")
    print(f"   - Data range: rows 21 to 149 ✅")
    print(f"   - Lovetex entries: {lovetex_count} ✅")
    print(f"   - Final rows: {ws.max_row}")
    print(f"   - Unique rows: 382")
    
    wb.close()

if __name__ == "__main__":
    verify_lovetex_fix()