#!/usr/bin/env python3
"""
Step 4: Article Information Filling - Content Preparation
SEDO TSS Converter Pipeline Step 4/6

LOGIC:
- Extract Article Name/Number from original input file (dynamic header detection)
- Search Article headers ONLY above "General Type" header to avoid conflicts
- Fill articles in R+ columns with merged cells (R1:R9), 90° text rotation
- Apply light orange background (FFD4B3) for article identification
- Place article numbers in row 10 for alignment with data structure

PIPELINE POSITION: Fourth step - adds article info before data transformation
INPUT: Original input file (data/input/input-X.xlsx) + Template (data/output/output-X-Step3.xlsx)
OUTPUT: Article-filled template (data/output/output-X-Step4.xlsx)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from typing import Union, Optional, List, Tuple
import argparse
import sys
import re
import shutil

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ArticleFiller:
    """
    Article Information Filler for Step 4 - Content Preparation
    
    Advanced article extraction and placement:
    - Dynamic header detection above "General Type" to avoid conflicts
    - Extracts multiple articles from original input file
    - Places articles in R+ columns (R, S, T, U...) with professional formatting
    - Creates merged cells with 90° text rotation and light orange background
    - Positions article numbers in row 10 for data structure alignment
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "data" / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def fill_article_info(self, input_file: Union[str, Path], 
                          step3_file: Union[str, Path],
                          output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Fill Article Name and Number from input file into Step 3 template
        
        Args:
            input_file: Original input file (input-X.xlsx) 
            step3_file: Step 3 template file (output-X-Step3.xlsx)
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to filled file
        """
        logger.info("📋 Step 4: Article Information Filling - Content Preparation")
        
        input_path = Path(input_file)
        step3_path = Path(step3_file)
        
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")
        if not step3_path.exists():
            raise FileNotFoundError(f"Step 3 file not found: {step3_path}")
        
        # Auto-generate output file if not provided
        if output_file is None:
            file_num = self._extract_file_number(step3_path.name)
            if file_num:
                output_file = self.output_dir / f"output-{file_num}-Step4.xlsx"
            else:
                base_name = step3_path.stem.replace('-Step3', '')
                output_file = self.output_dir / f"{base_name}-Step4.xlsx"
        else:
            output_file = Path(output_file)
        
        logger.info(f"Input File: {input_path}")
        logger.info(f"Step 3 Template: {step3_path}")
        logger.info(f"Output: {output_file}")
        
        # Load input file and find Article headers
        input_wb = openpyxl.load_workbook(str(input_path))
        input_ws = input_wb.active
        
        # Find Article Name and Number headers
        header_info = self._find_article_headers(input_ws)
        if not header_info:
            logger.warning("⚠️  No Article Name/Number headers found - creating output without article info")
            # Copy Step 3 file as-is
            shutil.copy2(str(step3_path), str(output_file))
            return str(output_file)
        
        name_col, no_col, header_row = header_info
        logger.info(f"Found headers: Article Name at col {name_col}, Article No. at col {no_col}, row {header_row}")
        
        # Extract article data
        articles = self._extract_article_data(input_ws, name_col, no_col, header_row)
        logger.info(f"Extracted {len(articles)} article(s)")
        
        # Copy Step 3 as starting point
        shutil.copy2(str(step3_path), str(output_file))
        output_wb = openpyxl.load_workbook(str(output_file))
        output_ws = output_wb.active
        
        # Fill article information into output template
        self._fill_output_template(output_ws, articles)
        
        # Save result
        try:
            output_wb.save(str(output_file))
            logger.info(f"✅ Step 4 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        return str(output_file)
    
    def _find_article_headers(self, worksheet) -> Optional[Tuple[int, int, int]]:
        """
        Find Article Name and Number header positions
        
        Search strategy:
        1. First find "General Type/Sub-Type in Connect" header row
        2. Search Article headers ONLY ABOVE that row
        
        Returns:
            (name_col, no_col, header_row) or None if not found
        """
        # Step 1: Find General Type/Sub-Type header row
        general_header_row = None
        for row in range(1, 50):  # Search in larger range
            for col in range(1, 16):
                cell_value = worksheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower().strip()
                    if 'general type' in cell_lower and ('sub-type' in cell_lower or 'connect' in cell_lower):
                        general_header_row = row
                        logger.debug(f"Found General Type header at row {row}")
                        break
            if general_header_row:
                break
        
        if not general_header_row:
            logger.warning("Could not find General Type/Sub-Type header - using default search range")
            general_header_row = 16  # Fallback
        
        # Step 2: Search Article headers ONLY ABOVE general header row
        name_col = None
        no_col = None
        header_row = None
        
        search_end_row = general_header_row  # Search from row 1 to general_header_row - 1
        logger.debug(f"Searching Article headers in rows 1 to {search_end_row - 1}")
        
        for row in range(1, search_end_row):
            for col in range(1, 16):
                cell_value = worksheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower().strip()
                    
                    if 'article name' in cell_lower:
                        name_col = col
                        header_row = row
                        logger.debug(f"Found Article Name at {chr(64+col) if col <= 26 else 'col'+str(col)}{row}")
                    elif 'article no' in cell_lower:
                        no_col = col
                        if header_row is None:  # Set header row if not set by name
                            header_row = row
                        logger.debug(f"Found Article No. at {chr(64+col) if col <= 26 else 'col'+str(col)}{row}")
        
        if name_col and no_col and header_row:
            logger.info(f"Found Article headers above General Type header: Name at col {name_col}, No. at col {no_col}, row {header_row}")
            return name_col, no_col, header_row
        else:
            if not name_col:
                logger.warning("Article Name header not found above General Type header")
            if not no_col:
                logger.warning("Article No. header not found above General Type header")
            return None
    
    def _extract_article_data(self, worksheet, name_col: int, no_col: int, header_row: int) -> List[Tuple[str, str]]:
        """
        Extract article data from columns below headers until empty
        
        Returns:
            List of (name, number) tuples
        """
        articles = []
        data_row = header_row + 1
        
        logger.debug(f"Extracting article data starting from row {data_row}")
        
        # Continue until both cells are empty
        while data_row <= worksheet.max_row:
            name_value = worksheet.cell(data_row, name_col).value
            no_value = worksheet.cell(data_row, no_col).value
            
            # Stop if both values are empty
            if not name_value and not no_value:
                break
            
            # Convert to strings, handle None values
            name_str = str(name_value) if name_value is not None else ""
            no_str = str(no_value) if no_value is not None else ""
            
            articles.append((name_str, no_str))
            logger.debug(f"Row {data_row}: Name=\"{name_str}\" | No.=\"{no_str}\"")
            
            data_row += 1
        
        return articles
    
    def _fill_output_template(self, worksheet, articles: List[Tuple[str, str]]):
        """
        Fill article information into output template
        
        New layout:
        - Article Numbers: Row 10, starting column R (R10, S10, T10, U10...)
        - Article Names: Merged cells R1:R9, S1:S9, T1:T9, U1:U9... with text rotation up
        """
        logger.debug(f"Filling {len(articles)} articles into output template")
        
        for i, (name, number) in enumerate(articles):
            # Calculate target column starting from R (column 18)
            target_col = 18 + i  # R=18, S=19, T=20, U=21, etc.
            col_letter = get_column_letter(target_col)
            
            # Fill article number in row 10 with very light orange background
            number_cell = worksheet.cell(10, target_col, number)
            number_cell.fill = PatternFill(start_color="00FFD4B3", end_color="00FFD4B3", fill_type="solid")  # Very light orange
            
            # Apply very light orange background to all cells before merging (rows 1-9)
            for row in range(1, 10):
                cell = worksheet.cell(row, target_col)
                cell.fill = PatternFill(start_color="00FFD4B3", end_color="00FFD4B3", fill_type="solid")  # Very light orange
            
            # Create merged range for article name (rows 1-9)
            merge_range = f"{col_letter}1:{col_letter}9"
            try:
                worksheet.merge_cells(merge_range)
            except Exception as e:
                logger.warning(f"Could not merge cells {merge_range}: {e}")
            
            # Fill article name in merged cell with rotation
            name_cell = worksheet.cell(1, target_col, name)
            name_cell.alignment = Alignment(
                textRotation=90,  # Rotate text up
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            # Reapply very light orange background to main cell after merge
            name_cell.fill = PatternFill(start_color="00FFD4B3", end_color="00FFD4B3", fill_type="solid")
            
            logger.debug(f"Article {i+1}: {col_letter}1:9=\"{name}\" (rotated) | {col_letter}10=\"{number}\"")
    
    def _extract_file_number(self, filename: str) -> str:
        """Extract file number from filename like 'output-1-Step3.xlsx'"""
        match = re.search(r'output-(\d+)', filename)
        return match.group(1) if match else ""
    
    def process_multiple_files(self, input_patterns: list, output_dir: Optional[str] = None) -> list:
        """
        Process multiple files
        
        Args:
            input_patterns: List of input file patterns or paths
            output_dir: Output directory (if None, use default)
            
        Returns:
            List of output file paths
        """
        if output_dir:
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(parents=True, exist_ok=True)
        
        results = []
        
        for pattern in input_patterns:
            # Handle glob patterns
            if '*' in str(pattern):
                input_files = list(self.base_dir.glob(str(pattern)))
            else:
                input_files = [Path(pattern)]
            
            for input_file in input_files:
                if input_file.exists() and input_file.suffix.lower() in ['.xlsx', '.xls']:
                    try:
                        # Find corresponding Step 3 file
                        file_num = self._extract_file_number(input_file.name)
                        if not file_num:
                            # Try to extract from input-X format
                            match = re.search(r'input-(\d+)', input_file.name)
                            file_num = match.group(1) if match else ""
                        
                        if file_num:
                            step3_file = self.base_dir / "data" / "output" / f"output-{file_num}-Step3.xlsx"
                        else:
                            logger.error(f"❌ Could not determine file number for: {input_file}")
                            continue
                        
                        if not step3_file.exists():
                            logger.error(f"❌ Step 3 file not found: {step3_file}")
                            continue
                        
                        result = self.fill_article_info(input_file, step3_file)
                        results.append(result)
                        logger.info(f"✅ Processed: {input_file} + {step3_file} → {result}")
                    except Exception as e:
                        logger.error(f"❌ Failed to process {input_file}: {e}")
                else:
                    logger.warning(f"⚠️  Skipped: {input_file} (not found or not Excel file)")
        
        return results

def main():
    """Command line interface for standalone article filling"""
    parser = argparse.ArgumentParser(description='Article Filler Step 4 - Standalone')
    parser.add_argument('input', nargs='*', help='Input file(s) or patterns (input-X.xlsx). If not provided, uses data/input/input-*.xlsx')
    parser.add_argument('--step3-file', help='Specific Step 3 file (only for single file mode)')
    parser.add_argument('-o', '--output', help='Output file or directory')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    parser.add_argument('--batch', action='store_true', help='Batch mode for multiple files')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize filler
    filler = ArticleFiller(args.base_dir)
    
    try:
        if args.batch or len(args.input) > 1 or not args.step3_file:
            # Multiple files mode or auto-detect mode
            input_patterns = args.input if args.input else ["data/input/input-*.xlsx"]
            output_dir = args.output if args.output else None
            results = filler.process_multiple_files(input_patterns, output_dir)
            
            print("\n📊 Batch Processing Results:")
            print(f"✅ Successfully processed: {len(results)} files")
            for result in results:
                print(f"   📁 {result}")
                
        else:
            # Single file mode - need both input and Step 3 files
            if not args.input:
                print("❌ Error: Must provide input file in single file mode")
                sys.exit(1)
            
            input_file = args.input[0]
            step3_file = args.step3_file
            
            result = filler.fill_article_info(input_file, step3_file, args.output)
            print(f"\n✅ Success!")
            print(f"📁 Output: {result}")
            
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()