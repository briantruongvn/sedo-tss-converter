#!/usr/bin/env python3
"""
Step 3: Template Creation - Structured Foundation
SEDO TSS Converter Pipeline Step 3/6

LOGIC:
- Create structured output template with 17 predefined headers (A-Q)
- Apply color-coded formatting: red (materials), blue (regulations), green (testing)
- Set column widths and alignment for optimal readability
- Position headers in row 10 to allow space for Article information above

PIPELINE POSITION: Third step - creates foundation template for data mapping
INPUT: Header-processed file (data/output/output-X-Step2.xlsx)
OUTPUT: Formatted template file (data/output/output-X-Step3.xlsx)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from pathlib import Path
from typing import Union, Optional
import argparse
import sys
import re
from validation_utils import ValidationError, handle_validation_error
from pipeline_validator import validate_before_pipeline
from pipeline_config import PipelineConfig

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class TemplateCreator:
    """
    Template Creator for Step 3 - Structured Foundation
    
    Creates formatted output template with:
    - Row 10: 17 column headers (A-Q) with color-coding and formatting
    - Optimized column widths for data display
    - Color scheme: Red (materials), Blue (regulations), Green (testing)
    - Rows 1-9: Reserved for Article information (filled by Step 4)
    """
    
    @classmethod
    def get_metadata(cls):
        """Get step metadata from centralized configuration"""
        return PipelineConfig.get_step(3)
    
    @property
    def step_name(self):
        """Get step display name"""
        return self.get_metadata().display_name
    
    @property
    def step_description(self):
        """Get step description"""
        return self.get_metadata().description
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "data" / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Template structure with formatting (background + font color + width)
        self.template_headers = [
            {"name": "Combination", "bg_color": "00FFFF00", "font_color": "00000000", "width": 15.0},
            {"name": "General Type Component(Type)", "bg_color": "00FF0000", "font_color": "00FFFFFF", "width": 20.0},
            {"name": "Sub-Type Component Identity Process Name", "bg_color": "00FF0000", "font_color": "00FFFFFF", "width": 25.0},
            {"name": "Material Designation", "bg_color": "00FF0000", "font_color": "00FFFFFF", "width": 18.0},
            {"name": "Material Distributor", "bg_color": "00FF0000", "font_color": "00FFFFFF", "width": 15.0},
            {"name": "Producer", "bg_color": "00FF0000", "font_color": "00FFFFFF", "width": 12.0},
            {"name": "Material Type In Process", "bg_color": "00FF0000", "font_color": "00FFFFFF", "width": 20.0},
            {"name": "Document type", "bg_color": "000000FF", "font_color": "00FFFFFF", "width": 15.0},
            {"name": "Requirement Source/TED", "bg_color": "000000FF", "font_color": "00FFFFFF", "width": 20.0},
            {"name": "Sub-type", "bg_color": "000000FF", "font_color": "00FFFFFF", "width": 12.0},
            {"name": "Regulation or substances", "bg_color": "000000FF", "font_color": "00FFFFFF", "width": 20.0},
            {"name": "Limit", "bg_color": "00B8E6B8", "font_color": "00000000", "width": 10.0},
            {"name": "Test method", "bg_color": "00B8E6B8", "font_color": "00000000", "width": 15.0},
            {"name": "Frequency", "bg_color": "00B8E6B8", "font_color": "00000000", "width": 12.0},
            {"name": "Level", "bg_color": "000000FF", "font_color": "00FFFFFF", "width": 10.0},
            {"name": "Warning Limit", "bg_color": "00B8E6B8", "font_color": "00000000", "width": 15.0},
            {"name": "Additional Information", "bg_color": "00B8E6B8", "font_color": "00000000", "width": 20.0}
        ]
        
        # Define styles
        self.row1_2_style = {
            "font": Font(bold=True, color="00000000"),
            "fill": PatternFill(start_color="00B8E6B8", end_color="00B8E6B8", fill_type="solid"),
            "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True)
        }
        
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def create_template(self, input_file: Union[str, Path], 
                       output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Create output template from Step 2 processed file
        
        Args:
            input_file: Input file from Step 2 (output-X-Step2.xlsx)
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to template file
        """
        logger.info("📋 Step 3: Create Output Template")
        
        input_path = Path(input_file)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")
        
        # Auto-generate output file if not provided
        if output_file is None:
            file_num = self._extract_file_number(input_path.name)
            if file_num:
                output_file = self.output_dir / f"output-{file_num}-Step3.xlsx"
            else:
                base_name = input_path.stem.replace('-Step2', '')
                output_file = self.output_dir / f"{base_name}-Step3.xlsx"
        else:
            output_file = Path(output_file)
        
        logger.info(f"Input: {input_path}")
        logger.info(f"Output: {output_file}")
        
        # Create new workbook with template structure
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Output Template"
        
        # Article name and number labels removed as requested
        
        # Row 10: Headers (17 columns A-Q) with specific formatting and column widths
        for col_idx, header_info in enumerate(self.template_headers, 1):
            cell = ws.cell(10, col_idx, header_info["name"])
            
            # Apply font with specific color for each column
            cell.font = Font(bold=True, color=header_info["font_color"])
            
            # Apply background color
            cell.fill = PatternFill(start_color=header_info["bg_color"], 
                                   end_color=header_info["bg_color"], 
                                   fill_type="solid")
            
            # Apply alignment
            cell.alignment = self.header_alignment
            
            # Set column width
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = header_info["width"]
        
        logger.info(f"✅ Created formatted template with {len(self.template_headers)} headers")
        
        # Save template
        try:
            wb.save(str(output_file))
            logger.info(f"✅ Step 3 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        return str(output_file)
    
    def _extract_file_number(self, filename: str) -> str:
        """Extract file number from filename like 'output-1-Step2.xlsx'"""
        match = re.search(r'output-(\d+)', filename)
        return match.group(1) if match else ""
    
    def create_multiple_templates(self, input_patterns: list, output_dir: Optional[str] = None) -> list:
        """
        Create templates for multiple files matching patterns
        
        Args:
            input_patterns: List of file patterns or paths
            output_dir: Output directory (if None, use default)
            
        Returns:
            List of output file paths
        """
        if output_dir:
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(parents=True, exist_ok=True)
        
        results = []
        
        for pattern in input_patterns:
            # Handle glob patterns
            if '*' in str(pattern):
                input_files = list(self.base_dir.glob(str(pattern)))
            else:
                input_files = [Path(pattern)]
            
            for input_file in input_files:
                if input_file.exists() and input_file.suffix.lower() in ['.xlsx', '.xls']:
                    try:
                        result = self.create_template(input_file)
                        results.append(result)
                        logger.info(f"✅ Processed: {input_file} → {result}")
                    except Exception as e:
                        logger.error(f"❌ Failed to process {input_file}: {e}")
                else:
                    logger.warning(f"⚠️  Skipped: {input_file} (not found or not Excel file)")
        
        return results

def main():
    """Command line interface for standalone template creation"""
    parser = argparse.ArgumentParser(description='Template Creator Step 3 - Standalone')
    parser.add_argument('input', nargs='+', help='Input file(s) or patterns (output-X-Step2.xlsx)')
    parser.add_argument('-o', '--output', help='Output file or directory')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    parser.add_argument('--batch', action='store_true', help='Batch mode for multiple files')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize creator
    creator = TemplateCreator(args.base_dir)
    
    try:
        if args.batch or len(args.input) > 1:
            # Multiple files mode
            output_dir = args.output if args.output else None
            results = creator.create_multiple_templates(args.input, output_dir)
            
            print("\n📊 Batch Processing Results:")
            print(f"✅ Successfully processed: {len(results)} files")
            for result in results:
                print(f"   📁 {result}")
                
        else:
            # Single file mode
            input_file = args.input[0]
            output_file = args.output
            
            result = creator.create_template(input_file, output_file)
            print(f"\n✅ Success!")
            print(f"📁 Output: {result}")
            
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()