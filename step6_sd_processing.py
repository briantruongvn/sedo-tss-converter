#!/usr/bin/env python3
"""
Step 6: SD Processing and De-duplication - Final Integration
SEDO TSS Converter Pipeline Step 6/6

LOGIC:
- Extract SD data from Step 2 "Requirements" row and below
- Parse multi-line G column values (SD specifications) and expand into separate rows
- Map SD data: A→B, B→C, E→F, F→D, G-lines→Q with "TR" document type
- Append SD rows to Step 5 output while preserving existing data
- Perform comprehensive de-duplication using ALL columns comparison
- Trim trailing spaces to avoid false duplicates

PIPELINE POSITION: Final step - completes data integration with SD processing
INPUT: Header-processed data (output-X-Step2.xlsx) + Populated template (output-X-Step5.xlsx)
OUTPUT: Complete processed file (output-X-Step6.xlsx) - FINAL RESULT
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from pathlib import Path
from typing import Union, Optional, List, Tuple
import argparse
import sys
import re
import shutil

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class SDProcessor:
    """
    SD Processor for Step 6 - Final Integration
    
    Comprehensive SD data processing and integration:
    - Locates "Requirements" section in Step 2 for SD data extraction
    - Validates and filters SD values (excludes "N/A", "Không")
    - Parses multi-line G column specifications into individual entries
    - Maps SD data to proper output columns with consistent formatting
    - Appends to existing Step 5 data without overwriting
    - Performs intelligent de-duplication across ALL columns
    - Produces final, clean, database-ready output
    """
    
    def __init__(self, base_dir: Optional[str] = None):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.output_dir = self.base_dir / "data" / "output"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def process_sd_data(self, step2_file: Union[str, Path], 
                       step4_file: Union[str, Path],
                       output_file: Optional[Union[str, Path]] = None) -> str:
        """
        Process SD data from Step 2 and append to Step 4
        
        Args:
            step2_file: Data source file (output-X-Step2.xlsx)
            step4_file: Template file (output-X-Step4.xlsx)  
            output_file: Optional output file path (if None, auto-generate)
            
        Returns:
            Path to processed file
        """
        logger.info("📋 Step 6: SD Processing and De-duplication")
        
        step2_path = Path(step2_file)
        step4_path = Path(step4_file)
        
        if not step2_path.exists():
            raise FileNotFoundError(f"Step 2 file not found: {step2_path}")
        if not step4_path.exists():
            raise FileNotFoundError(f"Step 4 file not found: {step4_path}")
        
        # Auto-generate output file if not provided
        if output_file is None:
            file_num = self._extract_file_number(step2_path.name)
            if file_num:
                output_file = self.output_dir / f"output-{file_num}-Step6.xlsx"
            else:
                base_name = step2_path.stem.replace('-Step2', '')
                output_file = self.output_dir / f"{base_name}-Step6.xlsx"
        else:
            output_file = Path(output_file)
        
        logger.info(f"Data Source (Step 2): {step2_path}")
        logger.info(f"Template (Step 5): {step4_path}")
        logger.info(f"Output: {output_file}")
        
        # Load source files
        step2_wb = openpyxl.load_workbook(str(step2_path))
        step2_ws = step2_wb.active
        
        # Copy Step 5 as starting point
        shutil.copy2(str(step4_path), str(output_file))
        output_wb = openpyxl.load_workbook(str(output_file))
        output_ws = output_wb.active
        
        # Step 1: Find requirements row in Step 2
        requirements_row = self._find_requirements_row(step2_ws)
        if requirements_row is None:
            raise ValueError("Could not find 'Requirements' row in Step 2 file")
        
        data_start_row = requirements_row + 1
        last_data_row = self._find_last_data_row(step2_ws, data_start_row)
        logger.info(f"Found requirements at row {requirements_row}, data from row {data_start_row} to {last_data_row}")
        
        # Step 2: Process SD data and append to Step 5
        current_output_row = output_ws.max_row + 1
        sd_rows_added = self._process_sd_rows(step2_ws, output_ws, data_start_row, last_data_row, current_output_row)
        
        logger.info(f"✅ Added {sd_rows_added} SD rows")
        
        # Step 3: Comprehensive de-duplication
        final_rows = self._deduplicate_rows(output_ws)
        logger.info(f"✅ After de-duplication: {final_rows} unique rows")
        
        # Save result
        try:
            output_wb.save(str(output_file))
            logger.info(f"✅ Step 6 completed: {output_file}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise
        
        return str(output_file)
    
    def _find_requirements_row(self, worksheet) -> Optional[int]:
        """Find row containing 'Requirements' in column A (case-insensitive)"""
        for row in range(1, worksheet.max_row + 1):
            cell_value = worksheet.cell(row, 1).value  # Column A
            if cell_value and 'requirement' in str(cell_value).lower():
                logger.debug(f"Found 'Requirements' row at row {row}")
                return row
        return None
    
    def _find_last_data_row(self, worksheet, start_row: int) -> int:
        """Find the last row that contains data in columns A or B (consistent with Step 4)"""
        max_row = worksheet.max_row
        
        # Search backwards from max_row to find last row with data in A or B
        for row in range(max_row, start_row - 1, -1):
            for col in range(1, 3):  # Columns A-B
                cell_value = worksheet.cell(row, col).value
                if cell_value and str(cell_value).strip():
                    return row
        
        return start_row  # Fallback to start_row if no data found
    
    def _process_sd_rows(self, source_ws, output_ws, start_row: int, end_row: int, output_start_row: int) -> int:
        """
        Process SD rows from Step 2 and append to Step 5 output
        
        Logic:
        - For each row from start_row to end_row
        - Check if G column has valid SD data (not N/A, not "Không")
        - Parse multi-line G column values
        - Create output rows with mapping: A19→B, B19→C, E19→F, F19→D, H19→P, G19→Q
        - Duplicate rows for each line in G column
        
        Returns:
            Number of rows added
        """
        rows_added = 0
        current_output_row = output_start_row
        
        for source_row in range(start_row, end_row + 1):
            # Get source data
            a_value = source_ws.cell(source_row, 1).value  # A
            b_value = source_ws.cell(source_row, 2).value  # B
            e_value = source_ws.cell(source_row, 5).value  # E
            f_value = source_ws.cell(source_row, 6).value  # F
            g_value = source_ws.cell(source_row, 7).value  # G
            h_value = source_ws.cell(source_row, 8).value  # H
            
            # Check if G column has valid data (skip N/A, "Không", empty)
            if not self._is_valid_sd_value(g_value):
                logger.debug(f"Skipping row {source_row}: Invalid G value '{g_value}'")
                continue
            
            # Parse multi-line G column value
            g_lines = self._parse_multiline_value(g_value)
            logger.debug(f"Row {source_row}: F='{f_value}', H='{h_value}', G has {len(g_lines)} lines")
            
            # Create output row(s) for each G line
            for line_idx, g_line in enumerate(g_lines):
                # Copy base formatting from previous row if not first
                if current_output_row > 11:
                    self._copy_row_formatting(output_ws, current_output_row - 1, current_output_row)
                
                # Map source values to output columns
                # A19 → B (dynamic row)
                output_ws.cell(current_output_row, 2, a_value)  # B
                # B19 → C
                output_ws.cell(current_output_row, 3, b_value)  # C
                # E19 → F  
                output_ws.cell(current_output_row, 6, e_value)  # F
                # F19 → D
                output_ws.cell(current_output_row, 4, f_value)  # D
                # H19 → P
                output_ws.cell(current_output_row, 16, h_value)  # P (column 16)
                # G19 lines → Q
                output_ws.cell(current_output_row, 17, g_line.strip())  # Q (column 17)
                
                logger.debug(f"Added output row {current_output_row}: F={f_value}, P={h_value}, Q='{g_line.strip()}'")
                current_output_row += 1
                rows_added += 1
        
        return rows_added
    
    def _is_valid_sd_value(self, value) -> bool:
        """
        Check if G column value is valid for SD processing
        
        Returns False for:
        - None/empty values
        - "N/A" (case-insensitive)
        - "Không" (Vietnamese for "No")
        - Whitespace-only strings
        """
        if not value:
            return False
        
        value_str = str(value).strip()
        if not value_str:
            return False
        
        # Check for invalid values
        invalid_values = ['n/a', 'không']
        if value_str.lower() in invalid_values:
            return False
        
        return True
    
    def _parse_multiline_value(self, value) -> List[str]:
        """
        Parse multi-line value from G column
        
        Examples:
        - "1/ SD MAT10 Jensberg (Nantong), printing" → ["1/ SD MAT10 Jensberg (Nantong), printing"]
        - "1/ SD MAT10\\n2/ SD MAT0054" → ["1/ SD MAT10", "2/ SD MAT0054"]
        """
        if not value:
            return []
        
        value_str = str(value).strip()
        if not value_str:
            return []
        
        # Split by newlines and filter out empty lines
        lines = [line.strip() for line in value_str.split('\n') if line.strip()]
        
        # If no newlines found, return the whole value as single line
        if len(lines) <= 1:
            return [value_str]
        
        return lines
    
    def _copy_row_formatting(self, worksheet, source_row: int, target_row: int):
        """Copy formatting from source row to target row for columns A-Q"""
        for col in range(1, 18):  # Columns A-Q
            source_cell = worksheet.cell(source_row, col)
            target_cell = worksheet.cell(target_row, col)
            
            # Copy formatting
            if source_cell.font:
                target_cell.font = Font(
                    bold=source_cell.font.bold,
                    color=source_cell.font.color
                )
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color,
                    fill_type=source_cell.fill.fill_type
                )
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )
    
    def _trim_trailing_spaces(self, worksheet):
        """Trim trailing spaces in all cells to avoid false duplicates"""
        logger.debug("Trimming trailing spaces in all cells...")
        
        for row in range(11, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row, col)
                if cell.value and isinstance(cell.value, str):
                    trimmed_value = cell.value.rstrip()
                    if trimmed_value != cell.value:
                        cell.value = trimmed_value
    
    def _deduplicate_rows(self, worksheet) -> int:
        """
        Remove duplicate rows based on ALL columns comparison
        
        Trim trailing spaces before comparison to avoid false duplicates
        Compare ALL columns to identify duplicates
        Keep first occurrence, remove subsequent duplicates
        
        Returns:
            Number of unique rows remaining
        """
        logger.info("🔄 Trimming trailing spaces and de-duplicating rows...")
        
        # First pass: trim trailing spaces in all cells
        self._trim_trailing_spaces(worksheet)
        
        # Collect all rows with their complete data (starting from row 11, after headers)
        rows_data = []
        max_col = worksheet.max_column
        
        for row in range(11, worksheet.max_row + 1):
            # Create a tuple of ALL columns for comparison
            row_data = tuple(worksheet.cell(row, col).value for col in range(1, max_col + 1))
            
            # Only consider rows with actual data (at least one non-None value)
            if any(val for val in row_data if val is not None):
                rows_data.append((row, row_data))
        
        # Find duplicates
        seen_rows = set()
        rows_to_delete = []
        
        for row_num, row_data in rows_data:
            if row_data in seen_rows:
                rows_to_delete.append(row_num)
                logger.debug(f"Marking row {row_num} for deletion (duplicate - all columns match)")
            else:
                seen_rows.add(row_data)
        
        # Delete duplicate rows (in reverse order to maintain row numbers)
        for row_num in sorted(rows_to_delete, reverse=True):
            worksheet.delete_rows(row_num, 1)
            logger.debug(f"Deleted duplicate row {row_num}")
        
        logger.info(f"🗑️  Removed {len(rows_to_delete)} duplicate rows")
        
        # Return final row count
        final_rows = worksheet.max_row
        return final_rows
    
    def _extract_file_number(self, filename: str) -> str:
        """Extract file number from filename like 'output-1-Step2.xlsx'"""
        match = re.search(r'output-(\d+)', filename)
        return match.group(1) if match else ""
    
    def process_multiple_files(self, step2_patterns: list, output_dir: Optional[str] = None) -> list:
        """
        Process multiple Step 2 files with their corresponding Step 4 files
        
        Args:
            step2_patterns: List of Step 2 file patterns or paths
            output_dir: Output directory (if None, use default)
            
        Returns:
            List of output file paths
        """
        if output_dir:
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(parents=True, exist_ok=True)
        
        results = []
        
        for pattern in step2_patterns:
            # Handle glob patterns
            if '*' in str(pattern):
                step2_files = list(self.base_dir.glob(str(pattern)))
            else:
                step2_files = [Path(pattern)]
            
            for step2_file in step2_files:
                if step2_file.exists() and step2_file.suffix.lower() in ['.xlsx', '.xls']:
                    try:
                        # Find corresponding Step 5 file
                        file_num = self._extract_file_number(step2_file.name)
                        if file_num:
                            step5_file = self.base_dir / "data" / "output" / f"output-{file_num}-Step5.xlsx"
                        else:
                            base_name = step2_file.stem.replace('-Step2', '')
                            step5_file = self.base_dir / "data" / "output" / f"{base_name}-Step5.xlsx"
                        
                        if not step5_file.exists():
                            logger.error(f"❌ Step 5 file not found: {step5_file}")
                            continue
                        
                        result = self.process_sd_data(step2_file, step5_file)
                        results.append(result)
                        logger.info(f"✅ Processed: {step2_file} + {step5_file} → {result}")
                    except Exception as e:
                        logger.error(f"❌ Failed to process {step2_file}: {e}")
                else:
                    logger.warning(f"⚠️  Skipped: {step2_file} (not found or not Excel file)")
        
        return results

def main():
    """Command line interface for standalone SD processing"""
    parser = argparse.ArgumentParser(description='SD Processor Step 6 - Standalone')
    parser.add_argument('input', nargs='*', help='Input Step 2 file(s) or patterns (output-X-Step2.xlsx). If not provided, uses data/output/*-Step2.xlsx')
    parser.add_argument('--step5-file', help='Specific Step 5 file (only for single file mode)')
    parser.add_argument('-o', '--output', help='Output file or directory')
    parser.add_argument('-d', '--base-dir', help='Base directory', default='.')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose logging')
    parser.add_argument('--batch', action='store_true', help='Batch mode for multiple files')
    
    args = parser.parse_args()
    
    # Configure logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Initialize processor
    processor = SDProcessor(args.base_dir)
    
    try:
        if args.batch or len(args.input) > 1 or not args.step5_file:
            # Multiple files mode or auto-detect mode
            input_patterns = args.input if args.input else ["data/output/*-Step2.xlsx"]
            output_dir = args.output if args.output else None
            results = processor.process_multiple_files(input_patterns, output_dir)
            
            print("\n📊 Batch Processing Results:")
            print(f"✅ Successfully processed: {len(results)} files")
            for result in results:
                print(f"   📁 {result}")
                
        else:
            # Single file mode - need both Step2 and Step5 files
            if not args.input:
                print("❌ Error: Must provide Step 2 file in single file mode")
                sys.exit(1)
            
            step2_file = args.input[0]
            step5_file = args.step5_file
            
            result = processor.process_sd_data(step2_file, step5_file, args.output)
            print(f"\n✅ Success!")
            print(f"📁 Output: {result}")
            
    except Exception as e:
        logger.error(f"❌ Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()